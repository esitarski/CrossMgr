#!/usr/bin/env python
#------------------------------------------------------------------------------	
# UltraClient.py: Ultra simulator program for testing Ultra interface and CrossMgr.
#
# Copyright (C) Edward Sitarski, 2012.
import re
import time
import math
import socket
import random
import datetime
now = datetime.datetime.now
import threading
from queue import Queue, Empty
from openpyxl.workbook import Workbook

#------------------------------------------------------------------------------	
# CrossMgr's port and socket.
from Ultra import DEFAULT_PORT
DEFAULT_HOST = '127.0.0.1'

#------------------------------------------------------------------------------	
from Ultra import EOL

#------------------------------------------------------------------------------	
# Create some random rider numbers.
random.seed( 10101010 )
seen = set()
nums = []
for i in range(25):
	while 1:
		x = random.randint(1,200)
		if x not in seen:
			seen.add( x )
			nums.append( x )
			break

#------------------------------------------------------------------------------	
# Create a Ultra-style hex tag for each number.
tag = dict( (n, '413A%02X' % n) for n in nums )
tag[random.choice(list(tag.keys()))] = 'E2001018860B01290700D0D8'
tag[random.choice(list(tag.keys()))] = 'E2001018860B01530700D138'
tag[random.choice(list(tag.keys()))] = 'E2001018860B01370700D0F8'
tag[random.choice(list(tag.keys()))] = '1'
tag[random.choice(list(tag.keys()))] = '2'

#------------------------------------------------------------------------
def getRandomData( starters = 64 ):
	firstNames = '''
1. Noah
2. Liam
3. Jacob
4. Mason
5. William
6. Ethan
7. Michael
8. Alexander
9. Jayden
10. Daniel
11. Elijah
12. Aiden
13. James
14. Benjamin
15. Matthew
16. Jackson
17. Logan
18. David
19. Anthony
20. Joseph
21. Joshua
22. Andrew
23. Lucas
24. Gabriel
25. Samuel
26. Christopher
27. John
28. Dylan
29. Isaac
30. Ryan
31. Nathan
32. Carter
33. Caleb
34. Luke
35. Christian
36. Hunter
37. Henry
38. Owen
39. Landon
40. Jack
41. Wyatt
42. Jonathan
43. Eli
44. Isaiah
45. Sebastian
46. Jaxon
47. Julian
48. Brayden
49. Gavin
50. Levi
51. Aaron
52. Oliver
53. Jordan
54. Nicholas
55. Evan
56. Connor
57. Charles
58. Jeremiah
59. Cameron
60. Adrian
61. Thomas
62. Robert
63. Tyler
64. Colton
65. Austin
66. Jace
67. Angel
68. Dominic
69. Josiah
70. Brandon
71. Ayden
72. Kevin
73. Zachary
74. Parker
75. Blake
76. Jose
77. Chase
78. Grayson
79. Jason
80. Ian
81. Bentley
82. Adam
83. Xavier
84. Cooper
85. Justin
86. Nolan
87. Hudson
88. Easton
89. Jase
90. Carson
91. Nathaniel
92. Jaxson
93. Kayden
94. Brody
95. Lincoln
96. Luis
97. Tristan
98. Damian
99. Camden
100. Juan
'''

	lastNames = '''
1. Smith  
2. Johnson  
3. Williams  
4. Jones  
5. Brown  
6. Davis  
7. Miller  
8. Wilson  
9. Moore  
10. Taylor  
11. Anderson  
12. Thomas  
13. Jackson  
14. White  
15. Harris  
16. Martin  
17. Thompson  
18. Garcia  
19. Martinez  
20. Robinson  
21. Clark  
22. Rodriguez  
23. Lewis  
24. Lee  
25. Walker  
26. Hall  
27. Allen  
28. Young  
29. Hernandez  
30. King  
31. Wright  
32. Lopez  
33. Hill  
34. Scott  
35. Green  
36. Adams  
37. Baker  
38. Gonzalez  
39. Nelson  
40. Carter  
41. Mitchell  
42. Perez  
43. Roberts  
44. Turner  
45. Phillips  
46. Campbell  
47. Parker  
48. Evans  
49. Edwards  
50. Collins  
51. Stewart  
52. Sanchez  
53. Morris  
54. Rogers  
55. Reed  
56. Cook  
57. Morgan  
58. Bell  
59. Murphy  
60. Bailey  
61. Rivera  
62. Cooper  
63. Richardson  
64. Cox  
65. Howard  
66. Ward  
67. Torres  
68. Peterson  
69. Gray  
70. Ramirez  
71. James  
72. Watson  
73. Brooks  
74. Kelly  
75. Sanders  
76. Price  
77. Bennett  
78. Wood  
79. Barnes  
80. Ross  
81. Henderson  
82. Coleman  
83. Jenkins  
84. Perry  
85. Powell  
86. Long  
87. Patterson  
88. Hughes  
89. Flores  
90. Washington  
91. Butler  
92. Simmons  
93. Foster  
94. Gonzales  
95. Bryant   
96. Alexander  
97. Russell  
98. Griffin   
99. Diaz  
'''

	teams = '''
The Cyclomaniacs
Pesky Peddlers
Geared Up
Spoke & Mirrors
The Handlebar Army
Wheels of Steel
The Chaingang
Saddled & Addled
The Cyclepaths
Tour de Farce
Old Cranks
Magically Bikelicious
Look Ma No Hands!
Pedal Pushers
Kicking Asphault
Velociposse
Road Rascals
Spin Doctors
'''

	firstNames = [line.split('.')[1].strip() for line in firstNames.split('\n') if line.strip()]
	lastNames = [line.split('.')[1].strip() for line in lastNames.split('\n') if line.strip()]
	teams = [line.strip() for line in teams.split('\n') if line.strip()]
	bibs = list( range( 1, 1+starters ) )
	
	random.shuffle( firstNames )
	random.shuffle( lastNames )
	random.shuffle( teams )
	random.shuffle( bibs )
	
	for i in range(starters):
		yield bibs[i], firstNames[i%len(firstNames)], lastNames[i&len(lastNames)], teams[i%len(teams)]
		
#------------------------------------------------------------------------------	
# Write out as a .xlsx file with the number tag data.
#
wb = Workbook()
ws = wb.create_sheet( "UltraTest", 0 )
for col, label in enumerate('Bib#,LastName,FirstName,Team,Tag'.split(',')):
	ws.cell( row = 1, column = col+1 ).value = label
rdata = [d for d in getRandomData(len(tag))]
rowCur = 2
for r, (n, t) in enumerate(tag.items()):
	if t in ('1', '2'):
		continue
	
	bib, firstName, lastName, Team = rdata[r]
	for c, v in enumerate([n, lastName, firstName, Team, t], 1):
		ws.cell( row = rowCur, column = c ).value = v
	rowCur += 1
wb.save('UltraTest.xlsx')
wb = None

#------------------------------------------------------------------------------	
# Also write out as a .csv file.
#
with open('UltraTest.csv', 'w') as f:
	f.write( 'Bib#,Tag,dummy3,dummy4,dummy5\n' )
	for n in nums:
		f.write( f'{n},{tag[n]}\n' )

sendDate = True

#------------------------------------------------------------------------------	
# Function to format number, lap and time in Ultra format
t1980 = datetime.datetime( 1980, 1, 1 )
messageCount = 0
def formatMessage( n, lap, t ):
	global messageCount
	messageCount += 1
	
	f, secs = math.modf( (t - t1980).total_seconds() )
	milliseconds = int(f*1000)
	
	message = "{},{},{},{},1,-8,0,1,1,1,456789,{}{}".format(
				0,
				tag[n],									# Tag code
				int(secs), milliseconds,
				messageCount,
				EOL,
			)
	return message

#------------------------------------------------------------------------------	
# Generate some random lap times.
random.seed()
numLapTimes = []
mean = 60.0							# Average lap time.
varFactor = 9.0 * 4.0
var = mean / varFactor				# Variance between riders.
lapMax = 6
for n in nums:
	lapTime = random.normalvariate( mean, mean/(varFactor * 4.0) )
	for lap in range(0, lapMax+1):
		numLapTimes.append( (n, lap, lapTime*lap) )
numLapTimes.sort( key = lambda x: (x[1], x[2]) )	# Sort by lap, then race time.

print( 'len(numLapTimes)=', len(numLapTimes) )

dBase = now()
dBase -= datetime.timedelta( seconds = 13*60+13.13 )	# Send the wrong time for testing purposes.
#------------------------------------------------------------------------------	
passingQ = Queue()

def sendData():
	iPassing = 1
	for iPassing in range(1,len(numLapTimes)):
		n, lap, t = numLapTimes[iPassing]
		dt = t - numLapTimes[iPassing-1][2]
		time.sleep( dt )
		passing = formatMessage( n, lap, dBase + datetime.timedelta(seconds = t) )
		passingQ.put( passing )
		# sys.stdout.write( 'passing: {}: {}\n'.format(len(passings), passing[:-len(EOL)]) )
		
thread = threading.Thread( target=sendData )
thread.daemon = True
thread.start()

serversocket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
serversocket.bind((DEFAULT_HOST, DEFAULT_PORT))
serversocket.listen(5)
while 1:
	(clientsocket, address) = serversocket.accept()
	clientsocket.settimeout( 2 )
	
	def sendResponse( response ):
		print( response )
		clientsocket.sendall( bytes(response + EOL) )
	
	print( 'Connection from:', address )

	active = False
	lastVoltage = now()
	while 1:
		try:
			message = clientsocket.recv( 4096 ).strip()
		except socket.timeout:
			if active:
				while 1:
					try:
						passing = passingQ.get_nowait()
						sendResponse( passing )
						passingQ.task_done()
					except Empty:
						break
			if (now() - lastVoltage).total_seconds() >= 10:
				sendResponse( 'V=25.0000' )
				lastVoltage = now()
			continue
		except Exception:
			break
		
		if not message:
			break
			
		print( 'message:', message )
		
		if message.startswith('t'):
			hh, mm, ss, day, month, year = [int(f) for f in re.split( '[^0-9]', message.split(' ', 1)[1] ) ]
			dBase = datetime.datetime( year, month, day, hh, mm, ss )
			sendResponse( dBase.strftime('%H:%M:%S %d-%m-%Y') )
			
		elif message.startswith('R'):
			active = True
			
		elif message.startswith('S'):
			active = False
		
		else:
			print( 'unknown:', message )

