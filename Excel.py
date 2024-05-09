import os
import re
import math
import xlrd
import openpyxl
import datetime
import unicodedata
import xml.etree.ElementTree
from mmap import mmap, ACCESS_READ

def toAscii( s ):
	if s is None:
		return ''
		
	if isinstance( s, bytes ):
		ret = s.encode()
	elif isinstance( s, str ):
		ret = unicodedata.normalize('NFKD', s).encode('ascii','ignore').decode()
	else:
		ret = str( s )
	
	if ret.endswith( '.0' ):
		ret = ret[:-2]
	return ret

def timeFromStr( value ):
	if not isinstance(value, str):
		return value
		
	# Convert times with AM/PM to 24-hour format.
	# First check that we have an am/pm time.
	# This will accept formats like 9:00a, 09:00am, 09:00 A, 09:00 AM, etc.
	v = value.strip().lower()
	if not (':' in v) or not ('a' in v or 'p' in v) or not v[0].isdigit() or not re.match('^[0-9: apm]+$', v):
		return value
	
	isPM = ('p' in v)			# Check for pm.
	fields = re.sub( '[^0-9:]', '', v ).split(':')
	while len(fields) < 3:		# Fix missing seconds.
		fields.append( '0' )
	if len(fields) > 3:			# Ignore extra fields.
		fields = fields[:3]
	hh, mm, ss = [int(f or 0) for f in fields]
	if isPM and hh != 12:		# Adjust hours to 24-hour format.
		hh += 12
	return '{:02d}:{:02d}:{:02d}'.format( hh, mm, ss )

#----------------------------------------------------------------------------

class ReadExcelXls:
	def __init__(self, filename):
		if not os.path.isfile(filename):
			raise ValueError( "{} is not a valid filename".format(filename) )
		with open(filename,'rb') as f:
			self.book = xlrd.open_workbook(
				filename=filename,
				file_contents=mmap(f.fileno(),0,access=ACCESS_READ),
			)
		# self.book = xlrd.open_workbook(filename)
		
	def is_nonempty_row(self, sheet, i):
		values = sheet.row_values(i)
		if isinstance(values[0], str) and values[0].startswith('#'):
			return False # ignorable comment row
		return any( bool(v) for v in values )
	
	def sheet_names( self ):
		return self.book.sheet_names()
		
	def _parse_row(self, sheet, row_index, date_as_tuple):
		""" Sanitize incoming excel data """
		# Data Type Codes:
		#  EMPTY 0
		#  TEXT 1 a Unicode string
		#  NUMBER 2 float
		#  DATE 3 float
		#  BOOLEAN 4 int; 1 means TRUE, 0 means FALSE
		#  ERROR 5
		values = []
		for type, value in zip(sheet.row_types(row_index), sheet.row_values(row_index)):
			if type == 2:
				try:
					if value == int(value):
						value = int(value)
				except TypeError:
					pass
			elif type == 3:
				if isinstance(value, float) and value < 1.0:
					t = value * (24.0*60.0*60.0)
					fract, secs = math.modf( t )
					if fract >= 0.99999:
						secs += 1.0
						fract = 0.0
					elif fract <= 0.00001:
						fract = 0.0
					
					secs = int(secs)
					if fract:
						value = '{:02d}:{:02d}:{:02d}.{}'.format(
							secs // (60*60), (secs // 60) % 60, secs % 60,
							'{:.20f}'.format(fract)[2:],
						)
					else:
						value = '{:02d}:{:02d}:{:02d}'.format(secs // (60*60), (secs // 60) % 60, secs % 60)
				else:
					try:
						datetuple = xlrd.xldate_as_tuple(value, self.book.datemode)
						validDate = True
					except Exception:
						value = 'UnreadableDate'
						validDate = False
					if validDate:
						if date_as_tuple:
							value = datetuple
						else:
							# time only - no date component
							if datetuple[0] == 0 and datetuple[1] == 0 and  datetuple[2] == 0:
								value = "{:02d}:{02d}:{02d}".format( *datetuple[3:] )
							# date only, no time
							elif datetuple[3] == 0 and datetuple[4] == 0 and datetuple[5] == 0:
								value = "{:%04d}/{:02d}/{:02d}".format( *datetuple[:3] )
							else: # full date
								value = "{:04d}/{:02d}/{:02d} {:02d}:{:02d}:{:02d}".format( *datetuple )
			elif type == 5:
				value = xlrd.error_text_from_code[value]
			values.append(timeFromStr(value))
		return values
		
	def iter_list(self, sname, date_as_tuple=False):
		sheet = self.book.sheet_by_name(sname) # XLRDError
		for i in range(sheet.nrows):
			yield self._parse_row(sheet, i, date_as_tuple)

#----------------------------------------------------------------------------

class ReadExcelXlsx:
	def __init__(self, filename):
		if not os.path.isfile(filename):
			raise ValueError( "{} is not a valid filename".format(filename) )
		self.book = openpyxl.load_workbook( filename, data_only=True )
		
	def is_nonempty_row(self, sheet, i):
		values = sheet.row_values(i)
		if isinstance(values[0], str) and values[0].startswith('#'):
			return False # ignorable comment row
		return any( bool(v) for v in values )
	
	def sheet_names( self ):
		return self.book.sheetnames
		
	def _parse_row(self, row, date_as_tuple):
		""" Sanitize incoming excel data """
		values = []
		for cell in row:
			value = cell.value
			if cell.data_type == 'n':
				try:
					if value == int(value):
						value = int(value)
				except TypeError:
					pass
			elif cell.data_type == 'd':
				if date_as_tuple:
					value = value.timetuple()
				else:
					if isinstance(value, datetime.time):
						value = value.strftime( '%H:%M:%S.%f' )
					elif isinstance(value, datetime.date):
						value = value.strftime( '%Y/%m/%d' )
					elif isinstance(value, datetime.datetime):
						value = value.strftime( '%Y/%m/%d %H:%M:%S.%f' )
					if isinstance(value, str) and value.endswith( '.000000' ):
						value = value[:-7]				
			values.append( timeFromStr(value) )
		return values
		
	def iter_list(self, sname, date_as_tuple=False):
		sheet = self.book[sname]
		for i, row in enumerate(sheet.iter_rows()):
			yield self._parse_row(row, date_as_tuple)

#----------------------------------------------------------------------------

def GetExcelReader( filename ):
	if filename.endswith( '.xls' ):
		return ReadExcelXls( filename )
	elif filename.endswith( '.xlsx' ) or filename.endswith( '.xlsm' ):
		return ReadExcelXlsx( filename )
	else:
		raise ValueError( '{} is not a recognized Excel format'.format(filename) )

#----------------------------------------------------------------------------

if __name__ == '__main__':
	#r = GetExcelReader( 'RaceResultTest.xlsx' )
	r = GetExcelReader( 'test2.xlsx' )
	for r, row in enumerate(r.iter_list(r.sheet_names()[0])):
		print( r, row )

