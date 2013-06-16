import wx
from wx.lib.wordwrap import wordwrap
import wx.lib.imagebrowser as imagebrowser
import sys
import os
import re
import datetime
import random
import time
import copy
import bisect
import json
import multiprocessing
import threading
import webbrowser
import zipfile
import base64
import cgi
import locale
locale.setlocale(locale.LC_ALL, '')
try:
	localDateFormat = locale.nl_langinfo( locale.D_FMT )
	localTimeFormat = locale.nl_langinfo( locale.T_FMT )
except:
	localDateFormat = '%b %d, %Y'
	localTimeFormat = '%I:%M%p'
	
import cPickle as pickle
from optparse import OptionParser

from ForecastHistory	import ForecastHistory
from NumKeypad			import NumKeypad
from Actions			import Actions
from Gantt				import Gantt
from History			import History
from RiderDetail		import RiderDetail
from Results			import Results
from Categories			import Categories
from Properties			import Properties, PropertiesDialog, ChangeProperties
from Recommendations	import Recommendations
from RaceAnimation		import RaceAnimation, GetAnimationData
from Search				import SearchDialog
from FtpWriteFile		import realTimeFtpPublish
from SetAutoCorrect		import SetAutoCorrectDialog
from DNSManager			import DNSManagerDialog
from USACExport			import USACExport
from HelpSearch			import HelpSearchDialog
import Utils
from Utils				import logCall
import Model
import VersionMgr
import JChipSetup
import JChipImport
import JChip
from JChip import ChipReaderEvent, EVT_CHIP_READER 
import OrionImport
import AlienImport
import ImpinjImport
import OutputStreamer
import FtpWriteFile
import GpxImport
import cStringIO as StringIO
from Undo import undo
from setpriority import setpriority
from Printing			import CrossMgrPrintout, getRaceCategories, ChoosePrintCategoriesDialog
import xlwt
from ExportGrid			import ExportGrid
import SimulationLapTimes
import Version
from ReadSignOnSheet	import GetExcelLink, ResetExcelLinkCache, ExcelLink
from SetGraphic			import SetGraphicDialog
from GetResults			import GetCategoryDetails
from PhotoFinish		import ResetPhotoInfoCache, DeletePhotos, SetCameraState
from PhotoViewer		import PhotoViewerDialog
from ReadTTStartTimesSheet import ImportTTStartTimes
import VideoBuffer

import wx.lib.agw.advancedsplash as AS
import openpyxl

'''
# Monkey patch threading so we can see where each thread gets started.
import types
import traceback
threading_start = threading.Thread.start
def loggingThreadStart( self, *args, **kwargs ):
	threading_start( self, *args, **kwargs )
	print self
	traceback.print_stack()
	print '----------------------------------'
threading.Thread.start = types.MethodType(loggingThreadStart, None, threading.Thread)
'''
#----------------------------------------------------------------------------------
		
def ShowSplashScreen():
	bitmap = wx.Bitmap( os.path.join(Utils.getImageFolder(), 'CrossMgrSplash.png'), wx.BITMAP_TYPE_PNG )
	
	# Write in the version number into the bitmap.
	w, h = bitmap.GetSize()
	dc = wx.MemoryDC()
	dc.SelectObject( bitmap )
	dc.SetFont( wx.FontFromPixelSize( wx.Size(0,h//10), wx.FONTFAMILY_SWISS, wx.NORMAL, wx.FONTWEIGHT_NORMAL ) )
	dc.DrawText( Version.AppVerName.replace('CrossMgr','Version'), w // 20, int(h * 0.44) )
	dc.SelectObject( wx.NullBitmap )
	
	# Show the splash screen.
	estyle = AS.AS_TIMEOUT | AS.AS_CENTER_ON_PARENT | AS.AS_SHADOW_BITMAP
	shadow = wx.Colour( 64, 64, 64 )
	showSeconds = 2.5
	try:
		frame = AS.AdvancedSplash(Utils.getMainWin(), bitmap=bitmap, timeout=int(showSeconds*1000),
								  extrastyle=estyle, shadowcolour=shadow)
	except:
		try:
			frame = AS.AdvancedSplash(Utils.getMainWin(), bitmap=bitmap, timeout=int(showSeconds*1000),
									  shadowcolour=shadow)
		except:
			pass
			
#----------------------------------------------------------------------------------
		
class MyTipProvider( wx.PyTipProvider ):
	def __init__( self, fname, tipNo = None ):
		self.tips = []
		try:
			with open(fname, 'r') as f:
				for line in f:
					line = line.strip()
					if line and line[0] != '#':
						self.tips.append( line )
			if tipNo is None:
				tipNo = (int(round(time.time() * 1000)) * 13) % (len(self.tips) - 1)
		except:
			pass
		if tipNo is None:
			tipNo = 0
		self.tipNo = tipNo
		wx.PyTipProvider.__init__( self, self.tipNo )
			
	def GetCurrentTip( self ):
		if self.tipNo < 0 or self.tipNo >= len(self.tips):
			self.tipNo = 0
		return self.tipNo
		
	def GetTip( self ):
		if not self.tips:
			return 'No tips available.'
		tip = self.tips[self.GetCurrentTip()].replace(r'\n','\n').replace(r'\t','    ')
		self.tipNo += 1
		return tip
		
	def PreprocessTip( self, tip ):
		return tip
		
	def DeleteFirstTip( self ):
		if self.tips:
			self.tips.pop(0)
		
	def __len__( self ):
		return len(self.tips)
		
	@property
	def CurrentTip( self ):
		return self.GetCurrentTip()
		
	@property
	def Tip( self ):
		return self.GetTip()

def ShowTipAtStartup():
	mainWin = Utils.getMainWin()
	if mainWin and not mainWin.config.ReadBool('showTipAtStartup', True):
		return
	
	tipFile = os.path.join(Utils.getImageFolder(), "tips.txt")
	try:
		provider = MyTipProvider( tipFile )
		'''
		if VersionMgr.isUpgradeRecommended():
			provider.tipNo = 0
		else:
			provider.DeleteFirstTip()
		'''
		showTipAtStartup = wx.ShowTip( None, provider, True )
		if mainWin:
			mainWin.config.WriteBool('showTipAtStartup', showTipAtStartup)
	except:
		pass

def replaceJsonVar( s, varName, value ):
	return s.replace( '%s = null' % varName, '%s = %s' % (varName, json.dumps(value)), 1 )

#----------------------------------------------------------------------------------
class CustomStatusBar(wx.StatusBar):
    def __init__(self, parent):
        wx.StatusBar.__init__(self, parent, -1)

        # This status bar has three fields
        self.SetFieldsCount(3)
        # Sets the three fields to be relative widths to each other.
        self.SetStatusWidths([-2, -1, -2])

        # Field 0 ... just text
        self.SetStatusText("A Custom StatusBar...", 0)

        # We're going to use a timer to drive a 'clock' in the last
        # field.
        self.timer = wx.PyTimer(self.Notify)
        self.timer.Start(1000)
        self.Notify()

    # Handles events from the timer we started in __init__().
    # We're using it to drive a 'clock' in field 2 (the third field).
    def Notify(self):
        t = time.localtime(time.time())
        st = time.strftime("%d-%b-%Y   %I:%M:%S", t)
        self.SetStatusText(st, 2)

#----------------------------------------------------------------------------------
		
class MainWin( wx.Frame ):
	def __init__( self, parent, id = wx.ID_ANY, title='', size=(200,200) ):
		wx.Frame.__init__(self, parent, id, title, size=size)

		Utils.setMainWin( self )

		# Add code to configure file history.
		self.filehistory = wx.FileHistory(8)
		self.config = wx.Config(appName="CrossMgr",
								vendorName="SmartCyclingSolutions",
								style=wx.CONFIG_USE_LOCAL_FILE)
		self.filehistory.Load(self.config)
		
		self.fileName = None
		self.numSelect = None
		
		#self.sb = CustomStatusBar( self )
		#self.SetStatusBar( self.sb )

		# Setup the objects for the race clock.
		self.timer = wx.Timer( self, id=wx.NewId() )
		self.secondCount = 0
		self.Bind( wx.EVT_TIMER, self.updateRaceClock, self.timer )

		self.simulateTimer = None
		self.simulateSeen = set()

		# Default print options.
		self.printData = wx.PrintData()
		self.printData.SetPaperId(wx.PAPER_LETTER)
		self.printData.SetPrintMode(wx.PRINT_MODE_PRINTER)
		self.printData.SetOrientation(wx.LANDSCAPE)

		# Configure the main menu.
		self.menuBar = wx.MenuBar(wx.MB_DOCKABLE)

		#-----------------------------------------------------------------------
		self.fileMenu = wx.Menu()

		self.fileMenu.Append( wx.ID_NEW , "&New...", "Create a new race" )
		self.Bind(wx.EVT_MENU, self.menuNew, id=wx.ID_NEW )

		idCur = wx.NewId()
		self.fileMenu.Append( idCur , "New Nex&t...", "Create a new race starting from the current race" )
		self.Bind(wx.EVT_MENU, self.menuNewNext, id=idCur )

		self.fileMenu.AppendSeparator()
		self.fileMenu.Append( wx.ID_OPEN , "&Open...", "Open a race" )
		self.Bind(wx.EVT_MENU, self.menuOpen, id=wx.ID_OPEN )

		idCur = wx.NewId()
		self.fileMenu.Append( idCur , "Open N&ext...", "Open the next race starting from the current race" )
		self.Bind(wx.EVT_MENU, self.menuOpenNext, id=idCur )

		self.fileMenu.AppendSeparator()
		idCur = wx.NewId()
		self.fileMenu.Append( idCur , "&Change Properties...", "Change the properties of the current race" )
		self.Bind(wx.EVT_MENU, self.menuChangeProperties, id=idCur )

		self.fileMenu.AppendSeparator()
		idCur = wx.NewId()
		self.fileMenu.Append( idCur , "&Restore from Original Input...", "Restore from Original Input" )
		self.Bind(wx.EVT_MENU, self.menuRestoreFromInput, id=idCur )

		self.fileMenu.AppendSeparator()
		self.fileMenu.Append( wx.ID_PAGE_SETUP , "Page &Setup...", "Setup the print page" )
		self.Bind(wx.EVT_MENU, self.menuPageSetup, id=wx.ID_PAGE_SETUP )

		self.fileMenu.Append( wx.ID_PREVIEW , "P&review Results...", "Preview the results on screen" )
		self.Bind(wx.EVT_MENU, self.menuPrintPreview, id=wx.ID_PREVIEW )

		self.fileMenu.Append( wx.ID_PRINT , "&Print Results...", "Print the results to a printer" )
		self.Bind(wx.EVT_MENU, self.menuPrint, id=wx.ID_PRINT )

		self.fileMenu.AppendSeparator()

		idCur = wx.NewId()
		self.fileMenu.Append( idCur , "&Publish Results as Excel...", "Publish Results as an Excel Spreadsheet (.xls)" )
		self.Bind(wx.EVT_MENU, self.menuPublishAsExcel, id=idCur )
		
		idCur = wx.NewId()
		self.fileMenu.Append( idCur , "Publish Results as &HTML...", "Publish Results as HTML (.html)" )
		self.Bind(wx.EVT_MENU, self.menuPublishHtmlRaceResults, id=idCur )

		self.fileMenu.AppendSeparator()
		
		idCur = wx.NewId()
		self.fileMenu.Append( idCur , "Publish HTML Results with FTP...", "Publish HTML Results to FTP" )
		self.Bind(wx.EVT_MENU, self.menuExportHtmlFtp, id=idCur )

		self.fileMenu.AppendSeparator()
		
		recent = wx.Menu()
		self.fileMenu.AppendMenu(wx.ID_ANY, "Recent Fil&es", recent)
		self.filehistory.UseMenu( recent )
		self.filehistory.AddFilesToMenu()
		
		self.fileMenu.AppendSeparator()

		self.fileMenu.Append( wx.ID_EXIT, "E&xit", "Exit CrossMan" )
		self.Bind(wx.EVT_MENU, self.menuExit, id=wx.ID_EXIT )
		
		self.Bind(wx.EVT_MENU_RANGE, self.menuFileHistory, id=wx.ID_FILE1, id2=wx.ID_FILE9)
		
		self.menuBar.Append( self.fileMenu, "&File" )

		#-----------------------------------------------------------------------
		self.editMenu = wx.Menu()
		self.undoMenuButton = wx.MenuItem( self.editMenu, wx.ID_UNDO , "&Undo\tCtrl+Z", "Undo the last edit" )
		img = wx.Image(os.path.join(Utils.getImageFolder(), 'Undo-icon.png'))
		self.undoMenuButton.SetBitmap( img.ConvertToBitmap(8) )
		self.editMenu.AppendItem( self.undoMenuButton )
		self.Bind(wx.EVT_MENU, self.menuUndo, id=wx.ID_UNDO )
		self.undoMenuButton.Enable( False )

		self.redoMenuButton = wx.MenuItem( self.editMenu, wx.ID_REDO , "&Redo\tCtrl+Y", "Redo the last edit" )
		img = wx.Image(os.path.join(Utils.getImageFolder(), 'Redo-icon.png'))
		self.redoMenuButton.SetBitmap( img.ConvertToBitmap(8) )
		self.editMenu.AppendItem( self.redoMenuButton )
		self.Bind(wx.EVT_MENU, self.menuRedo, id=wx.ID_REDO )
		self.redoMenuButton.Enable( False )
		self.editMenu.AppendSeparator()
		
		self.editMenu.Append( wx.ID_FIND, "&Find...\tCtrl+F", "Find a Rider" )
		self.Bind(wx.EVT_MENU, self.menuFind, id=wx.ID_FIND )
		
		self.editMenu.AppendSeparator()
		idCur = wx.NewId()
		self.editMenu.Append( idCur, '&Change "Autocorrect"...', 'Change "Autocorrect"...' )
		self.Bind( wx.EVT_MENU, self.menuAutocorrect, id=idCur )
		
		img = None
		self.editMenuItem = self.menuBar.Append( self.editMenu, "&Edit" )

		#-----------------------------------------------------------------------
		self.dataMgmtMenu = wx.Menu()
		
		idCur = wx.NewId()
		self.dataMgmtMenu.Append( idCur , "&Link to External Excel Data...", "Link to information in an Excel spreadsheet" )
		self.Bind(wx.EVT_MENU, self.menuLinkExcel, id=idCur )
		
		self.dataMgmtMenu.AppendSeparator()
		
		#-----------------------------------------------------------------------
		idCur = wx.NewId()
		self.dataMgmtMenu.Append( idCur, '&Add DNS from External Excel Data...', 'Add DNS...' )
		self.Bind( wx.EVT_MENU, self.menuDNS, id=idCur )
		
		self.dataMgmtMenu.AppendSeparator()
		
		#-----------------------------------------------------------------------
		categoryMgmtMenu = wx.Menu()
		self.dataMgmtMenu.AppendMenu(wx.ID_ANY, "Category Mgmt", categoryMgmtMenu)

		idCur = wx.NewId()
		categoryMgmtMenu.Append( idCur , "&Import Categories from File...", "Import Categories from File" )
		self.Bind(wx.EVT_MENU, self.menuImportCategories, id=idCur )

		idCur = wx.NewId()
		categoryMgmtMenu.Append( idCur , "&Export Categories to File...", "Export Categories to File" )
		self.Bind(wx.EVT_MENU, self.menuExportCategories, id=idCur )

		#-----------------------------------------------------------------------
		idCur = wx.NewId()
		self.dataMgmtMenu.Append( idCur , "Export History to Excel...", "Export History to Excel File" )
		self.Bind(wx.EVT_MENU, self.menuExportHistory, id=idCur )

		idCur = wx.NewId()
		self.dataMgmtMenu.Append( idCur , "Export Raw Data as &HTML...", "Export raw data as HTML (.html)" )
		self.Bind(wx.EVT_MENU, self.menuExportHtmlRawData, id=idCur )

		idCur = wx.NewId()
		self.dataMgmtMenu.Append( idCur , "Export Results in &USAC Excel Format...", "Export Results in USAC Excel Format" )
		self.Bind(wx.EVT_MENU, self.menuExportUSAC, id=idCur )

		self.dataMgmtMenu.AppendSeparator()
		idCur = wx.NewId()
		self.dataMgmtMenu.Append( idCur , "&Import Time Trial Start Times...", "Import Time Trial Start Times" )
		self.Bind(wx.EVT_MENU, self.menuImportTTStartTimes, id=idCur )
		
		self.dataMgmtMenu.AppendSeparator()
		idCur = wx.NewId()
		self.dataMgmtMenu.Append( idCur , "&Import Course in GPX format...", "Import Course in GPX format" )
		self.Bind(wx.EVT_MENU, self.menuImportGpx, id=idCur )
		
		idCur = wx.NewId()
		self.dataMgmtMenu.Append( idCur , "&Export Course as KMZ Virtual Tour...", "Export Course as KMZ Virtual Tour (Requires Google Earth)" )
		self.Bind(wx.EVT_MENU, self.menuExportCourseAsKml, id=idCur )
		
		self.menuBar.Append( self.dataMgmtMenu, "&DataMgmt" )

		#-----------------------------------------------------------------------

		# Configure the field of the display.

		# Forecast/History shown in left pane of scrolled window.
		forecastHistoryWidth = 265
		sty = wx.BORDER_SUNKEN
		self.splitter = wx.SplitterWindow( self )
		self.splitter.SetMinimumPaneSize( forecastHistoryWidth )
		self.forecastHistory = ForecastHistory( self.splitter, style=sty )

		# Other data shown in right pane.
		self.notebook		= wx.Notebook(	self.splitter, 1000, style=sty )
		self.notebook.Bind( wx.EVT_NOTEBOOK_PAGE_CHANGED, self.onPageChanging )
		
		# Add all the pages to the notebook.
		self.pages = []

		def addPage( page, name ):
			self.notebook.AddPage( page, name )
			self.pages.append( page )
			
		self.attrClassName = [
			[ 'actions',		Actions,			'Actions' ],
			[ 'record',			NumKeypad,			'Record' ],
			[ 'results',		Results,			'Results' ],
			[ 'history',		History,			'History' ],
			[ 'riderDetail',	RiderDetail,		'RiderDetail' ],
			[ 'gantt', 			Gantt,				'Chart' ],
			[ 'raceAnimation',	RaceAnimation,		'Animation' ],
			[ 'recommendations',Recommendations,	'Recommendations' ],
			[ 'categories', 	Categories,			'Categories' ],
			[ 'properties',		Properties,			'Properties' ]
		]
		
		for i, (a, c, n) in enumerate(self.attrClassName):
			setattr( self, a, c(self.notebook) )
			addPage( getattr(self, a), '%d. %s' % (i+1, n) )

		self.riderDetailDialog = None
		self.splitter.SplitVertically( self.forecastHistory, self.notebook, forecastHistoryWidth + 80)
		self.splitter.UpdateSize()

		#------------------------------------------------------------------------------
		# Create a menu for quick navigation
		self.pageMenu = wx.Menu()
		self.idPage = {}
		jumpToIds = []
		for i, p in enumerate(self.pages):
			name = self.notebook.GetPageText(i)
			idCur = wx.NewId()
			self.idPage[idCur] = i
			self.pageMenu.Append( idCur, '%s\tF%d' % (name, i+1), "Jump to %s page" % name )
			self.Bind(wx.EVT_MENU, self.menuShowPage, id=idCur )
			jumpToIds.append( idCur )
			
		self.menuBar.Append( self.pageMenu, "&JumpTo" )

		#-----------------------------------------------------------------------
		self.chipMenu = wx.Menu()

		idCur = wx.NewId()
		self.chipMenu.Append( idCur, "JChip &Setup...", "Configure and Test JChip Reader" )
		self.Bind(wx.EVT_MENU, self.menuJChip, id=idCur )
		
		chipImportMenu = wx.Menu()
		self.chipMenu.AppendMenu(wx.ID_ANY, "&Import", chipImportMenu)
		
		idCur = wx.NewId()
		chipImportMenu.Append( idCur , "JChip File...", "JChip Formatted File" )
		self.Bind(wx.EVT_MENU, self.menuJChipImport, id=idCur )
		
		idCur = wx.NewId()
		chipImportMenu.Append( idCur , "Alien File...", "Alien Formatted File" )
		self.Bind(wx.EVT_MENU, self.menuAlienImport, id=idCur )
		
		idCur = wx.NewId()
		chipImportMenu.Append( idCur , "Impinj File...", "Impinj Formatted File" )
		self.Bind(wx.EVT_MENU, self.menuImpinjImport, id=idCur )
		
		idCur = wx.NewId()
		chipImportMenu.Append( idCur , "Orion File...", "Orion Formatted File" )
		self.Bind(wx.EVT_MENU, self.menuOrionImport, id=idCur )
		
		self.menuBar.Append( self.chipMenu, "Chip&Reader" )

		#-----------------------------------------------------------------------
		self.demoMenu = wx.Menu()

		idCur = wx.NewId()
		self.demoMenu.Append( idCur , "&Simulate Race...", "Simulate a race" )
		self.Bind(wx.EVT_MENU, self.menuSimulate, id=idCur )

		self.menuBar.Append( self.demoMenu, "Dem&o" )

		#-----------------------------------------------------------------------
		self.optionsMenu = wx.Menu()
		idCur = wx.NewId()
		self.menuItemHighPrecisionTimes = self.optionsMenu.Append( idCur , "&Show 100s of a second", "Show 100s of a second", wx.ITEM_CHECK )
		self.Bind( wx.EVT_MENU, self.menuShowHighPrecisionTimes, id=idCur )
		
		idCur = wx.NewId()
		self.menuItemPlaySounds = self.optionsMenu.Append( idCur , "&Play Sounds", "Play Sounds", wx.ITEM_CHECK )
		self.playSounds = self.config.ReadBool('playSounds', True)
		self.menuItemPlaySounds.Check( self.playSounds )
		self.Bind( wx.EVT_MENU, self.menuPlaySounds, id=idCur )
		
		idCur = wx.NewId()
		self.menuItemSyncCategories = self.optionsMenu.Append( idCur , "Sync &Categories between Tabs", "Sync Categories between Tabs", wx.ITEM_CHECK )
		self.Bind( wx.EVT_MENU, self.menuSyncCategories, id=idCur )
		
		self.optionsMenu.AppendSeparator()
		
		idCur = wx.NewId()
		self.optionsMenu.Append( idCur , "Set Contact &Email...", "Set Contact Email for HTML Output" )
		self.Bind(wx.EVT_MENU, self.menuSetContactEmail, id=idCur )
		
		idCur = wx.NewId()
		self.optionsMenu.Append( idCur , "Set &Graphic...", "Set Graphic for Output" )
		self.Bind(wx.EVT_MENU, self.menuSetGraphic, id=idCur )
		
		self.optionsMenu.AppendSeparator()

		idCur = wx.NewId()
		self.optionsMenu.Append( idCur , "Copy Log File to &Clipboard", "Copy Log File to &Clipboard" )
		self.Bind(wx.EVT_MENU, self.menuCopyLogFileToClipboard, id=idCur )
		
		self.menuBar.Append( self.optionsMenu, "&Options" )
		
		#-----------------------------------------------------------------------
		self.helpMenu = wx.Menu()

		idCur = wx.NewId()
		self.helpMenu.Append( idCur , "&QuickStart...", "Get started with CrossMgr Now..." )
		self.Bind(wx.EVT_MENU, self.menuHelpQuickStart, id=idCur )
		idCur = wx.NewId()
		self.helpMenu.Append( idCur, "Help &Search...", "Search Help..." )
		self.Bind(wx.EVT_MENU, self.menuHelpSearch, id=idCur )
		self.helpSearch = HelpSearchDialog( self, wx.ID_ANY, title='Help Search' )
		self.helpMenu.Append( wx.ID_HELP, "&Help...", "Help about CrossMgr..." )
		self.Bind(wx.EVT_MENU, self.menuHelp, id=wx.ID_HELP )
		
		self.helpMenu.AppendSeparator()

		self.helpMenu.Append( wx.ID_ABOUT , "&About...", "About CrossMgr..." )
		self.Bind(wx.EVT_MENU, self.menuAbout, id=wx.ID_ABOUT )

		idCur = wx.NewId()
		self.helpMenu.Append( idCur , "&Tips at Startup...", "Enable/Disable Tips at Startup..." )
		self.Bind(wx.EVT_MENU, self.menuTipAtStartup, id=idCur )

		self.menuBar.Append( self.helpMenu, "&Help" )

		#------------------------------------------------------------------------------
		self.SetMenuBar( self.menuBar )

		#------------------------------------------------------------------------------
		# Set the accelerator table so we can switch windows with the function keys.
		accTable = [(wx.ACCEL_NORMAL, wx.WXK_F1 + i, jumpToIds[i]) for i in xrange(len(jumpToIds))]
		self.contextHelp = wx.NewId()
		self.Bind(wx.EVT_MENU, self.onContextHelp, id=self.contextHelp )
		accTable.append( (wx.ACCEL_CTRL, ord('H'), self.contextHelp) )
		accTable.append( (wx.ACCEL_SHIFT, wx.WXK_F1, self.contextHelp) )
		aTable = wx.AcceleratorTable( accTable )
		self.SetAcceleratorTable(aTable)
		
		#------------------------------------------------------------------------------
		self.Bind(wx.EVT_CLOSE, self.onCloseWindow)
		self.Bind(EVT_CHIP_READER, self.handleChipReaderEvent)
		self.lastPhotoTime = datetime.datetime.now()
		
		self.photoDialog = PhotoViewerDialog( self, wx.ID_ANY, "PhotoViewer", size=(600,400) )

	def handleChipReaderEvent( self, event ):
		race = Model.race
		if not race or not race.isRunning() or not getattr(race, 'enableUSBCamera', False):
			return
		if not getattr(race, 'tagNums', None):
			JChipSetup.GetTagNums()
		if not race.tagNums:
			return
		
		for tag, dt in event.tagTimes:
			if race.startTime > dt:
				continue
			try:
				num = race.tagNums[tag]
			except (TypeError, ValueError, KeyError):
				continue
			delta = dt - race.startTime
			t = delta.total_seconds()
			
			race.photoCount = getattr(race,'photoCount',0) + Utils.TakePhoto( num, t )
	
	def menuDNS( self, event ):
		dns = DNSManagerDialog( self )
		dns.ShowModal()
		dns.Destroy()
		
	def menuFind( self, event ):
		if not getattr(self, 'findDialog', None):
			self.findDialog = SearchDialog( self, wx.ID_ANY )
		self.findDialog.refresh()
		self.findDialog.Show()
		
	def menuUndo( self, event ):
		undo.doUndo()
		self.refresh()
		
	def menuRedo( self, event ):
		undo.doRedo()
		self.refresh()
		
	def menuAutocorrect( self, event ):
		undo.pushState()
		with Model.LockRace() as race:
			if not race:
				return
			categories = race.getCategoriesInUse()
			if not categories:
				return
		SetAutoCorrectDialog( self, categories ).ShowModal()
	
	def menuShowHighPrecisionTimes( self, event ):
		with Model.LockRace() as race:
			if race:
				race.highPrecisionTimes = self.menuItemHighPrecisionTimes.IsChecked()
		wx.CallAfter( self.refresh )
	
	def menuSyncCategories( self, event ):
		with Model.LockRace() as race:
			if race:
				race.syncCategories = self.menuItemSyncCategories.IsChecked()
	
	def menuPlaySounds( self, event ):
		self.playSounds = self.menuItemPlaySounds.IsChecked()
		self.config.WriteBool( 'playSounds', self.playSounds )
	
	def menuTipAtStartup( self, event ):
		showing = self.config.ReadBool('showTipAtStartup', True)
		if Utils.MessageOKCancel( self, 'Turn Off Tips at Startup?' if showing else 'Show Tips at Startup?', 'Tips at Startup' ):
			self.config.WriteBool( 'showTipAtStartup', showing ^ True )

	def menuRestoreFromInput( self, event ):
		if not Model.race:
			Utils.MessageOK(self, "You must have a valid race.", "No Valid Race", iconMask=wx.ICON_ERROR)
			return
		if not Utils.MessageOKCancel( self, "This will restore the race from the original input and replace your adds/splits/deletes.\nAre you sure you want to continue?",
									"Restore from Original Input", iconMask=wx.ICON_WARNING ):
			return
				
		startTime, endTime, numTimes = OutputStreamer.ReadStreamFile()
		if not numTimes:
			Utils.MessageOK( self, 'No data found.\n\nCheck "%s" file.' % OutputStreamer.getFileName(), "No Data Found" )
			return
		undo.pushState()
		with Model.LockRace() as race:
			race.clearAllRiderTimes()
			race.startTime = startTime
			race.endTime = endTime
			for num, t in numTimes:
				race.addTime( num, t )
			race.numLaps = race.getMaxLap()
			race.setChanged()
		wx.CallAfter( self.refresh )
			
	def menuChangeProperties( self, event ):
		if not Model.race:
			Utils.MessageOK(self, "You must have a valid race.", "No Valid Race", iconMask=wx.ICON_ERROR)
			return
		ChangeProperties( self )
		
	def menuJChip( self, event ):
		dlg = JChipSetup.JChipSetupDialog( self )
		dlg.ShowModal()
		dlg.Destroy()

	def menuJChipImport( self, event ):
		correct, reason = JChipSetup.CheckExcelLink()
		explain = 	'JChip Import requires that you have a valid Excel sheet with associated tags and Bib numbers.\n\n' \
					'See documentation for details.'
		if not correct:
			Utils.MessageOK( self, 'Problems with Excel sheet.\n\n    Reason: %s\n\n%s' % (reason, explain),
									title = 'Excel Link Problem', iconMask = wx.ICON_ERROR )
			return
			
		dlg = JChipImport.JChipImportDialog( self )
		dlg.ShowModal()
		dlg.Destroy()
		wx.CallAfter( self.refresh )
		
	def menuAlienImport( self, event ):
		correct, reason = JChipSetup.CheckExcelLink()
		explain = 	'Alien Import requires that you have a valid Excel sheet with associated tags and Bib numbers.\n\n' \
					'See documentation for details.'
		if not correct:
			Utils.MessageOK( self, 'Problems with Excel sheet.\n\n    Reason: %s\n\n%s' % (reason, explain),
									title = 'Excel Link Problem', iconMask = wx.ICON_ERROR )
			return
			
		dlg = AlienImport.AlienImportDialog( self )
		dlg.ShowModal()
		dlg.Destroy()
		wx.CallAfter( self.refresh )
		
	def menuImpinjImport( self, event ):
		correct, reason = JChipSetup.CheckExcelLink()
		explain = 	'Impinj Import requires that you have a valid Excel sheet with associated tags and Bib numbers.\n\n' \
					'See documentation for details.'
		if not correct:
			Utils.MessageOK( self, 'Problems with Excel sheet.\n\n    Reason: %s\n\n%s' % (reason, explain),
									title = 'Excel Link Problem', iconMask = wx.ICON_ERROR )
			return
			
		dlg = ImpinjImport.ImpinjImportDialog( self )
		dlg.ShowModal()
		dlg.Destroy()
		wx.CallAfter( self.refresh )
		
	def menuOrionImport( self, event ):
		correct, reason = JChipSetup.CheckExcelLink()
		explain = 	'Orion Import requires that you have a valid Excel sheet with associated tags and Bib numbers.\n\n' \
					'See documentation for details.'
		if not correct:
			Utils.MessageOK( self, 'Problems with Excel sheet.\n\n    Reason: %s\n\n%s' % (reason, explain),
									title = 'Excel Link Problem', iconMask = wx.ICON_ERROR )
			return
			
		dlg = OrionImport.OrionImportDialog( self )
		dlg.ShowModal()
		dlg.Destroy()
		wx.CallAfter( self.refresh )
		
	def menuShowPage( self, event ):
		self.showPage( self.idPage[event.GetId()] )
		
	def getDirName( self ):
		return Utils.getDirName()
		
	def menuSetContactEmail( self, event = None ):
		email = self.config.Read( 'email', 'my_name@my_address' )
		dlg = wx.TextEntryDialog( self, message='Contact Email:', caption='Contact Email for HTML output', defaultValue=email )
		result = dlg.ShowModal()
		if result == wx.ID_OK:
			value = dlg.GetValue()
			self.config.Write( 'email', value )
			self.config.Flush()
		dlg.Destroy()

	def menuSetGraphic( self, event ):
		imgPath = self.getGraphicFName()
		dlg = SetGraphicDialog( self, graphic = imgPath )
		if dlg.ShowModal() == wx.ID_OK:
			imgPath = dlg.GetValue()
			self.config.Write( 'graphic', imgPath )
			self.config.Flush()
		dlg.Destroy()
	
	def menuCopyLogFileToClipboard( self, event ):
		try:
			logData = open(redirectFileName).read()
		except IOError:
			Utils.MessageOK(self, "Unable to open log file.", "Error", wx.ICON_ERROR )
			return
			
		logData = logData.split( '\n' )
		logData = logData[-1000:]
		logData = '\n'.join( logData )
		
		dataObj = wx.TextDataObject()
		dataObj.SetText(logData)
		if wx.TheClipboard.Open():
			wx.TheClipboard.SetData( dataObj )
			wx.TheClipboard.Close()
			Utils.MessageOK(self, "Log file copied to clipboard.\nYou can now paste it into an email.", "Success" )
		else:
			Utils.MessageOK(self, "Unable to open the clipboard.", "Error", wx.ICON_ERROR )
	
	def getGraphicFName( self ):
		defaultFName = os.path.join(Utils.getImageFolder(), 'CrossMgrHeader.png')
		graphicFName = self.config.Read( 'graphic', defaultFName )
		if graphicFName != defaultFName:
			try:
				with open(graphicFName, 'rb') as f:
					return graphicFName
			except IOError:
				pass
		return defaultFName
	
	def getGraphicBase64( self ):
		graphicFName = self.getGraphicFName()
		if not graphicFName:
			return None
		fileType = os.path.splitext(graphicFName)[1].lower()
		if not fileType:
			return None
		fileType = fileType[1:]
		if fileType == 'jpg':
			fileType = 'jpeg'
		if fileType not in ['png', 'gif', 'jpeg']:
			return None
		try:
			with open(graphicFName, 'rb') as f:
				b64 = 'data:image/%s;base64,%s' % (fileType, base64.standard_b64encode(f.read()))
				return b64
		except IOError:
			pass
		return None
	
	def menuPageSetup( self, event ):
		psdd = wx.PageSetupDialogData(self.printData)
		psdd.CalculatePaperSizeFromId()
		dlg = wx.PageSetupDialog(self, psdd)
		dlg.ShowModal()

		# this makes a copy of the wx.PrintData instead of just saving
		# a reference to the one inside the PrintDialogData that will
		# be destroyed when the dialog is destroyed
		self.printData = wx.PrintData( dlg.GetPageSetupData().GetPrintData() )
		dlg.Destroy()

	def menuPrintPreview( self, event ):
		data = wx.PrintDialogData(self.printData)
		printout = CrossMgrPrintout()
		printout2 = CrossMgrPrintout()
		self.preview = wx.PrintPreview(printout, printout2, data)

		self.preview.SetZoom( 110 )
		if not self.preview.Ok():
			return

		pfrm = wx.PreviewFrame(self.preview, self, "Print preview")

		pfrm.Initialize()
		pfrm.SetPosition(self.GetPosition())
		pfrm.SetSize(self.GetSize())
		pfrm.Show(True)

	@logCall
	def menuPrint( self, event ):
		if not Model.race:
			return
		
		cpcd = ChoosePrintCategoriesDialog( self )
		x, y = self.GetPosition().Get()
		x += wx.SystemSettings.GetMetric(wx.SYS_FRAMESIZE_X, self)
		y += wx.SystemSettings.GetMetric(wx.SYS_FRAMESIZE_Y, self)
		cpcd.SetPosition( (x, y) )
		cpcd.SetSize( (450, 300) )
		result = cpcd.ShowModal()
		categories = cpcd.categories
		cpcd.Destroy()
		if not categories or result != wx.ID_OK:
			return
	
		self.printData.SetFilename( self.fileName if self.fileName else '' )
		pdd = wx.PrintDialogData(self.printData)
		pdd.SetAllPages( True )
		pdd.EnableSelection( False )
		pdd.EnablePageNumbers( False )
		pdd.EnableHelp( False )
		pdd.EnablePrintToFile( False )
		
		printer = wx.Printer(pdd)
		printout = CrossMgrPrintout( categories )

		if not printer.Print(self, printout, True):
			if printer.GetLastError() == wx.PRINTER_ERROR:
				Utils.MessageOK(self, "There was a printer problem.\nCheck your printer setup.", "Printer Error",iconMask=wx.ICON_ERROR)
		else:
			self.printData = wx.PrintData( printer.GetPrintDialogData().GetPrintData() )

		printout.Destroy()

	@logCall
	def menuLinkExcel( self, event = None ):
		if not Model.race:
			Utils.MessageOK(self, "You must have a valid race.", "Link ExcelSheet", iconMask=wx.ICON_ERROR)
			return
		self.showPageName( 'Results' )
		self.closeFindDialog()
		ResetExcelLinkCache()
		gel = GetExcelLink( self, getattr(Model.race, 'excelLink', None) )
		link = gel.show()
		undo.pushState()
		with Model.LockRace() as race:
			if not link:
				try:
					del race.excelLink
				except AttributeError:
					pass
			else:
				race.excelLink = link
			race.setChanged()
			race.resetAllCaches()
		self.writeRace()
		ResetExcelLinkCache()
		self.refresh()
		
	#--------------------------------------------------------------------------------------------

	@logCall
	def menuPublishAsExcel( self, event ):
		self.commit()
		if self.fileName is None or len(self.fileName) < 4:
			return

		xlFName = self.fileName[:-4] + '.xls'
		dlg = wx.DirDialog( self, 'Folder to write "%s"' % os.path.basename(xlFName),
						style=wx.DD_DEFAULT_STYLE, defaultPath=os.path.dirname(xlFName) )
		ret = dlg.ShowModal()
		dName = dlg.GetPath()
		dlg.Destroy()
		if ret != wx.ID_OK:
			return

		xlFName = os.path.join( dName, os.path.basename(xlFName) )

		wb = xlwt.Workbook()
		raceCategories = getRaceCategories()
		for catName, category in raceCategories:
			if catName == 'All' and len(raceCategories) > 1:
				continue
			sheetName = re.sub('[+!#$%&+~`".:;|\\/?*\[\] ]+', ' ', catName)
			sheetName = sheetName[:31]
			sheetCur = wb.add_sheet( sheetName )
			export = ExportGrid()
			export.setResultsOneList( category, showLapsFrequency = 1 )
			export.toExcelSheet( sheetCur )

		try:
			wb.save( xlFName )
			webbrowser.open( xlFName, new = 2, autoraise = True )
			Utils.MessageOK(self, 'Excel file written to:\n\n   %s' % xlFName, 'Excel Write')
		except IOError:
			Utils.MessageOK(self,
						'Cannot write "%s".\n\nCheck if this spreadsheet is open.\nIf so, close it, and try again.' % xlFName,
						'Excel File Error', iconMask=wx.ICON_ERROR )
						
	#--------------------------------------------------------------------------------------------
	def getEmail( self ):
		return self.config.Read('email', '')
	
	def addResultsToHtmlStr( self, html ):
		payload = {}
		payload['raceName'] = os.path.basename(self.fileName)[:-4]
			
		with Model.LockRace() as race:
			year, month, day = [int(v) for v in race.date.split('-')]
			timeComponents = [int(v) for v in race.scheduledStart.split(':')]
			if len(timeComponents) < 3:
				timeComponents.append( 0 )
			hour, minute, second = timeComponents
			raceTime = datetime.datetime( year, month, day, hour, minute, second )
			title = '%s Results for %s Start on %s' % ( race.name, raceTime.strftime(localTimeFormat), raceTime.strftime(localDateFormat) )
			html = html.replace( 'CrossMgr Race Results by Edward Sitarski', cgi.escape(title) )
			
			payload['organizer']		= getattr(race, 'organizer', '')
			payload['reverseDirection']	= getattr(race, 'reverseDirection', False)
			payload['finishTop']		= getattr(race, 'finishTop', False)
			payload['isTimeTrial']		= getattr(race, 'isTimeTrial', False)
			payload['rfid']				= getattr(race, 'enableJChipIntegration', False)
			payload['raceIsRunning']	= race.isRunning()
			payload['raceNotes']		= cgi.escape(getattr(race, 'notes', '')).replace('\n','{{br/}}')
			if race.startTime:
				raceStartTime = (race.startTime - race.startTime.replace( hour=0, minute=0, second=0 )).total_seconds()
				payload['raceStartTime']= raceStartTime
			tLastRaceTime = race.lastRaceTime()
			courseCoordinates, gpsPoints, gpsAltigraph, totalElevationGain = None, None, None, None
			geoTrack = getattr(race, 'geoTrack', None)
			if geoTrack is not None:
				courseCoordinates = geoTrack.asCoordinates()
				gpsPoints = geoTrack.asExportJson()
				gpsAltigraph = geoTrack.getAltigraph()
				totalElevationGain = geoTrack.totalElevationGainM
		
		tNow = datetime.datetime.now()
		payload['timestamp']			= [tNow.ctime(), tLastRaceTime]
		payload['email']				= self.getEmail()
		payload['data']					= GetAnimationData(getExternalData = True)
		payload['catDetails']			= GetCategoryDetails()
		payload['version']				= Version.AppVerName
		if gpsPoints:
			payload['gpsPoints']		= gpsPoints
		if courseCoordinates:
			payload['courseCoordinates'] = courseCoordinates
			# Fix the google maps template.
			templateFile = os.path.join(Utils.getHtmlFolder(), 'VirtualTourTemplate.html')
			try:
				with open(templateFile) as fp:
					template = fp.read()
				# Sanitize the template into a safe json string.
				template = template.replace( '<', '{{' ).replace( '>', '}}' )
				payload['virtualRideTemplate'] = template
			except:
				pass
		if totalElevationGain:
			payload['gpsTotalElevationGain'] = totalElevationGain
		if gpsAltigraph:
			payload['gpsAltigraph'] = gpsAltigraph

		html = replaceJsonVar( html, 'payload', payload )
		graphicBase64 = self.getGraphicBase64()
		if graphicBase64:
			try:
				iStart = html.index( 'src="data:image/png' )
				iEnd = html.index( '"/>', iStart )
				html = ''.join( [html[:iStart], 'src="%s"' % graphicBase64, html[iEnd+1:]] )
			except ValueError:
				pass
		return html
	
	@logCall
	def menuPublishHtmlRaceResults( self, event ):
		self.commit()
		if self.fileName is None or len(self.fileName) < 4:
			return
			
		if not self.getEmail():
			if Utils.MessageOKCancel( self,
				'Your Email contact is not set.\n\nConfigure it now?\n\n(you can always change it later from "Options|Set Contact Email...")',
				'Set Email Contact', wx.ICON_EXCLAMATION ):
				self.menuSetContactEmail()
	
		# Get the folder to write the html file.
		fname = self.fileName[:-4] + '.html'
		dlg = wx.DirDialog( self, 'Folder to write "%s"' % os.path.basename(fname),
							style=wx.DD_DEFAULT_STYLE, defaultPath=os.path.dirname(fname) )
		ret = dlg.ShowModal()
		dName = dlg.GetPath()
		dlg.Destroy()
		if ret != wx.ID_OK:
			return

		# Read the html template.
		htmlFile = os.path.join(Utils.getHtmlFolder(), 'RaceAnimation.html')
		try:
			with open(htmlFile) as fp:
				html = fp.read()
		except:
			Utils.MessageOK('Cannot read HTML template file.  Check program installation.',
							'Html Template Read Error', iconMask=wx.ICON_ERROR )
			return
			
		html = self.addResultsToHtmlStr( html )
			
		# Write out the results.
		fname = os.path.join( dName, os.path.basename(fname) )
		try:
			with open(fname, 'w') as fp:
				fp.write( html )
			webbrowser.open( fname, new = 0, autoraise = True )
			Utils.MessageOK(self, 'Html Race Animation written to:\n\n   %s' % fname, 'Html Write')
		except:
			Utils.MessageOK(self, 'Cannot write HTML file (%s).' % fname,
							'Html Write Error', iconMask=wx.ICON_ERROR )
	
	@logCall
	def menuExportHtmlFtp( self, event ):
		self.commit()
		if self.fileName is None or len(self.fileName) < 4:
			Utils.MessageOK(self, 'Ftp Upload Failed.  Error:\n\n    No race loaded.', 'Ftp Upload Failed', iconMask=wx.ICON_ERROR )
			return
	
		dlg = FtpWriteFile.FtpPublishDialog( self )
		ret = dlg.ShowModal()
		dlg.Destroy()
		if ret != wx.ID_OK:
			return
	
		# Read the html template.
		htmlFile = os.path.join(Utils.getHtmlFolder(), 'RaceAnimation.html')
		try:
			with open(htmlFile) as fp:
				html = fp.read()
		except:
			Utils.MessageOK('Cannot read HTML template file.  Check program installation.',
							'Html Template Read Error', iconMask=wx.ICON_ERROR )
			return
		
		html = self.addResultsToHtmlStr( html )
		
		# Publish the results using ftp.
		with Model.LockRace() as race:
			host		= getattr( race, 'ftpHost', '' )
			user		= getattr( race, 'ftpUser', '' )
			passwd		= getattr( race, 'ftpPassword', '' )
			serverPath	= getattr( race, 'ftpPath', '' )
			fname		= os.path.basename( self.fileName[:-4] + '.html' )
			file		= StringIO.StringIO( html )
			urlFull		= getattr( race, 'urlFull', '' )
		
		if not host:
			Utils.MessageOK(self, 'Ftp Upload Failed.  Error:\n\n    Missing host name.', 'Ftp Upload Failed', iconMask=wx.ICON_ERROR )
			return
		
		wx.BeginBusyCursor()
		try:
			FtpWriteFile.FtpWriteFile(	host		= host,
										user		= user,
										serverPath	= serverPath,
										passwd		= passwd,
										fname		= fname,
										file		= file )
			Utils.MessageOK(self, 'Ftp Upload Succeeded.', 'Ftp Upload Succeeded')
		except Exception, e:
			Utils.MessageOK(self, 'Ftp Upload Failed.  Error:\n\n%s' % str(e), 'Ftp Upload Failed', iconMask=wx.ICON_ERROR )
		wx.EndBusyCursor()
		
		# Automatically open the browser on the published file for testing.
		if urlFull:
			webbrowser.open( urlFull, new = 0, autoraise = True )
			
	#--------------------------------------------------------------------------------------------
	@logCall
	def menuImportTTStartTimes( self, event ):
		if self.fileName is None or len(self.fileName) < 4:
			return
		
		with Model.LockRace() as race:
			if not race:
				return
			if not race.isTimeTrial:
				Utils.MessageOK( self, 'You must set TimeTrial mode first.', 'Race must be TimeTrial' )
				return
			
		ImportTTStartTimes( self )
	
	@logCall
	def menuImportGpx( self, event ):
		if self.fileName is None or len(self.fileName) < 4:
			return
		
		with Model.LockRace() as race:
			if not race:
				return
			gt = GpxImport.GetGeoTrack( self, getattr(race, 'geoTrack', None), getattr(race, 'geoTrackFName', '') )
			
		geoTrack, geoTrackFName = gt.show()
		
		with Model.LockRace() as race:
			if not geoTrackFName:
				race.geoTrack, race.geoTrackFName = None, None
			else:
				race.geoTrack, race.geoTrackFName = geoTrack, geoTrackFName
				#with open('track.json', 'w') as f:
				#	f.write( json.dumps(race.geoTrack.asExportJson()) )
			race.showOval = (race.geoTrack is None)
			race.setChanged()
			
		self.showPageName( 'Animation' )
		self.refresh()
		
	@logCall
	def menuExportCourseAsKml( self, event ):
		with Model.LockRace() as race:
			if not race:
				return
				
			if not getattr(race, 'geoTrack', None):
				Utils.MessageOK( self, 'No GPX Course Loaded.\nNothing to export.', 'No GPX Course Loaded' )
				return
				
			geoTrack = race.geoTrack
			
			# Get the folder to write the html file.
			fname = self.fileName[:-4] + 'Course.kmz'
			dlg = wx.DirDialog( self, 'Folder to write "%s"' % os.path.basename(fname),
								style=wx.DD_DEFAULT_STYLE, defaultPath=os.path.dirname(fname) )
			ret = dlg.ShowModal()
			dName = dlg.GetPath()
			dlg.Destroy()
			if ret != wx.ID_OK:
				return
			
			fname = os.path.join( dName, os.path.basename(fname) )
			courseFName = os.path.splitext(os.path.basename(fname))[0] + '.kml'
			
			zf = zipfile.ZipFile( fname, 'w', zipfile.ZIP_DEFLATED )
			zf.writestr( courseFName, geoTrack.asKmlTour(race.name) )
			zf.close()
			
		webbrowser.open( fname, new = 0, autoraise = True )
		Utils.MessageOK(self, 'Course Virtual Tour written to KMZ file:\n\n   %s\n\nGoogle Earth Launched.' % fname, 'KMZ Write')
		
	@logCall
	def menuExportHtmlRawData( self, event ):
		self.commit()
		if self.fileName is None or len(self.fileName) < 4:
			return
		
		with Model.LockRace() as race:
			startTime, endTime, rawData = race.getRawData()
		
		if not rawData:
			Utils.MessageOK( self, 'Raw race data file:\n\n    "%s"\n\nis empty/missing.' % OutputStreamer.getFileName(),
					'Missing Raw Race Data', wx.ICON_ERROR )
			return
		
		# Get the folder to write the html file.
		fname = self.fileName[:-4] + 'RawData.html'
		dlg = wx.DirDialog( self, 'Folder to write "%s"' % os.path.basename(fname),
							style=wx.DD_DEFAULT_STYLE, defaultPath=os.path.dirname(fname) )
		ret = dlg.ShowModal()
		dName = dlg.GetPath()
		dlg.Destroy()
		if ret != wx.ID_OK:
			return

		# Read the html template.
		htmlFile = os.path.join(Utils.getHtmlFolder(), 'RawData.html')
		try:
			with open(htmlFile) as fp:
				html = fp.read()
		except:
			Utils.MessageOK('Cannot read HTML template file.  Check program installation.',
							'Html Template Read Error', iconMask=wx.ICON_ERROR )
			return
		
		# Replace parts of the file with the race information.
		html = replaceJsonVar( html, 'raceName', os.path.basename(self.fileName)[:-4] )
		html = replaceJsonVar( html, 'raceStart', (startTime - datetime.datetime.combine(startTime.date(), datetime.time())).total_seconds() )
		html = replaceJsonVar( html, 'rawData', rawData )
		
		with Model.LockRace() as race:
			try:
				externalFields = race.excelLink.getFields()
				externalInfo = race.excelLink.read()
			except:
				externalFields = []
				externalInfo = {}

			ignoreFields = set( ['Bib#', 'License'] )
			for f in ignoreFields:
				try:
					externalFields.remove( f )
				except ValueError:
					pass
			
			# Add the race category to the info.
			externalFields.insert( 0, 'Race Cat.' )
			if 'LastName' in externalFields or 'FirstName' in externalFields:
				externalFields.insert( 1, 'Name' )
				try:
					externalFields.remove( 'LastName' )
				except ValueError:
					pass
				try:
					externalFields.remove( 'FirstName' )
				except ValueError:
					pass
			
			seen = set()
			for d in rawData:
				num = d[1]
				if num not in seen:
					seen.add( num )
					category = race.getCategory( num )
					externalInfo.setdefault(num, {})['Race Cat.'] = category.name if category else 'Unknown'
					info = externalInfo[num]
					info['Name'] = Utils.CombineFirstLastName( info.get('FirstName', ''), info.get('LastName', '') )
			
			# Remove info that does not correspond to a rider in the race.
			for num in [n for n in externalInfo.iterkeys() if n not in seen]:
				del externalInfo[num]
			
			# Remove extra info fields.
			for num, info in externalInfo.iteritems():
				for f in ignoreFields:
					try:
						del info[f]
					except KeyError:
						pass
			
			html = replaceJsonVar( html, 'externalFields', externalFields )
			html = replaceJsonVar( html, 'externalInfo', externalInfo )
			
			year, month, day = [int(v) for v in race.date.split('-')]
			timeComponents = [int(v) for v in race.scheduledStart.split(':')]
			if len(timeComponents) < 3:
				timeComponents.append( 0 )
			hour, minute, second = timeComponents
			raceTime = datetime.datetime( year, month, day, hour, minute, second )
			title = '%s Raw Data for %s Start on %s' % ( race.name, raceTime.strftime(localTimeFormat), raceTime.strftime(localDateFormat) )
			html = html.replace( 'CrossMgr Race Results by Edward Sitarski', cgi.escape(title) )
			html = replaceJsonVar( html, 'organizer', getattr(race, 'organizer', '') )
			
		html = replaceJsonVar( html, 'timestamp', datetime.datetime.now().ctime() )
		
		graphicBase64 = self.getGraphicBase64()
		if graphicBase64:
			try:
				iStart = html.index( 'var imageSrc =' )
				iEnd = html.index( "';", iStart )
				html = ''.join( [html[:iStart], "var imageSrc = '%s';" % graphicBase64, html[iEnd+2:]] )
			except ValueError:
				pass
			
		# Write out the results.
		fname = os.path.join( dName, os.path.basename(fname) )
		try:
			with open(fname, 'w') as fp:
				fp.write( html )
			webbrowser.open( fname, new = 0, autoraise = True )
			Utils.MessageOK(self, 'Html Raw Data written to:\n\n   %s' % fname, 'Html Write')
		except:
			Utils.MessageOK(self, 'Cannot write HTML file (%s).' % fname,
							'Html Write Error', iconMask=wx.ICON_ERROR )
	
	#--------------------------------------------------------------------------------------------
	@logCall
	def onCloseWindow( self, event ):
		self.showPageName( 'Results' )
		with Model.LockRace() as race:
			if race is not None:
				race.resetAllCaches()
		self.writeRace()

		try:
			self.timer.Stop()
			del self.timer
		except AttributeError:
			pass

		try:
			self.simulateTimer.Stop()
			del self.simulateTimer
		except AttributeError:
			pass
		
		OutputStreamer.StopStreamer()
		VideoBuffer.Shutdown()
		JChip.Cleanuplistener()
		
		wx.Exit()

	def writeRace( self ):
		self.commit()
		with Model.LockRace() as race:
			if race is not None:
				with open(self.fileName, 'wb') as fp:
					pickle.dump( race, fp, 2 )
				race.setChanged( False )

	def setActiveCategories( self ):
		with Model.LockRace() as race:
			if race is None:
				return
			race.setActiveCategories()

	@logCall
	def menuNew( self, event ):
		self.showPageName( 'Actions' )
		self.closeFindDialog()
		
		race = Model.race
		if race:
			geoTrack, geoTrackFName = getattr(race, 'geoTrack', None), getattr(race, 'geoTrackFName', None)
			excelLink = getattr(race, 'excelLink', None)
		else:
			geoTrack, geoTrackFName, excelLink = None, None, None
			
		geoTrack, geoTrackFName = None, None		# Do not retain the GPX file after a full new.
		
		dlg = PropertiesDialog(self, -1, 'Configure Race', style=wx.DEFAULT_DIALOG_STYLE )
		ret = dlg.ShowModal()
		fileName = dlg.GetPath()
		categoriesFile = dlg.GetCategoriesFile()
		properties = dlg.properties

		if ret != wx.ID_OK:
			return
			
		# Check for existing file.
		if os.path.exists(fileName) and \
		   not Utils.MessageOKCancel(self, 'File "%s" already exists.  Overwrite?' % fileName, 'File Exists'):
			return

		# Try to open the file.
		try:
			with open(fileName, 'wb') as fp:
				pass
		except IOError:
			Utils.MessageOK( self, 'Cannot open "%s".' % fileName, 'Cannot Open File', iconMask=wx.ICON_ERROR )
			return

		with Model.LockRace() as race:
			if race:
				race.resetAllCaches()
		
		self.writeRace()
		
		# Create a new race and initialize it with the properties.
		self.fileName = fileName
		Model.resetCache()
		ResetExcelLinkCache()
		Model.setRace( Model.Race() )
		properties.update()
		ResetPhotoInfoCache( self.fileName )
		self.updateRecentFiles()

		importedCategories = False
		race = Model.race
		if categoriesFile:
			try:
				with open(categoriesFile, 'r') as fp:
					race.importCategories( fp )
				importedCategories = True
			except IOError:
				Utils.MessageOK( self, "Cannot open file:\n%s" % categoriesFile, "File Open Error", iconMask=wx.ICON_ERROR)
			except (ValueError, IndexError):
				Utils.MessageOK( self, "Bad file format:\n%s" % categoriesFile, "File Read Error", iconMask=wx.ICON_ERROR)

		# Create some defaults so the page is not blank.
		if not importedCategories:
			race.categoriesImportFile = ''
			race.setCategories( [{'name':'Category %d-%d'	% (max(1, i*100), (i+1)*100-1),
								  'catStr':'%d-%d'			% (max(1, i*100), (i+1)*100-1)} for i in xrange(8)] )
		else:
			race.categoriesImportFile = categoriesFile
			
		if geoTrack:
			race.geoTrack, race.geoTrackFName = geoTrack, geoTrackFName
			distance = geoTrack.length if race.distanceUnit == race.UnitKm else geoTrack.length * 0.621371
			if distance > 0.0:
				for c in race.categories.itervalues():
					c.distance = distance
			race.showOval = False
		if excelLink:
			race.excelLink = excelLink

		self.setNumSelect( None )
		self.writeRace()
		self.showPageName( 'Actions' )
		self.refreshAll()
	
	@logCall
	def menuNewNext( self, event ):
		race = Model.race
		if race is None:
			self.menuNew( event )
			return

		self.closeFindDialog()
		self.showPageName( 'Actions' )
		race.resetAllCaches()
		ResetExcelLinkCache()
		self.writeRace()
		
		# Save categories, gpx track and Excel link and use them in the next race.
		categoriesSave = race.categories
		geoTrack, geoTrackFName = getattr(race, 'geoTrack', None), getattr(race, 'geoTrackFName', None)
		excelLink = getattr(race, 'excelLink', None)
		race = None
		
		# Configure the next race.
		dlg = PropertiesDialog(self, -1, 'Configure Race', style=wx.DEFAULT_DIALOG_STYLE )
		dlg.properties.refresh()
		dlg.properties.incNext()
		dlg.properties.setEditable( True )
		dlg.folder.SetValue(os.path.dirname(self.fileName))
		dlg.properties.updateFileName()
		ret = dlg.ShowModal()
		fileName = dlg.GetPath()
		categoriesFile = dlg.GetCategoriesFile()
		properties = dlg.properties

		# Check if user canceled.
		if ret != wx.ID_OK:
			return

		# Check for existing file.
		if os.path.exists(fileName) and \
		   not Utils.MessageOKCancel(self, 'File "%s" already exists.  Overwrite?' % fileName, 'File Exists'):
			return

		# Try to open the file.
		try:
			with open(fileName, 'wb') as fp:
				pass
		except IOError:
			Utils.MessageOK(self, 'Cannot open "%s".' % fileName, 'Cannot Open File', iconMask=wx.ICON_ERROR )
			return

		# Create a new race and initialize it with the properties.
		self.fileName = fileName
		Model.resetCache()
		ResetExcelLinkCache()
		ResetPhotoInfoCache( self.fileName )
		
		# Save the current Ftp settings.
		ftpPublish = FtpWriteFile.FtpPublishDialog( self )

		Model.newRace()
		properties.update()			# Apply the new properties
		ftpPublish.setRaceAttr()	# Apply the ftp properties
		ftpPublish.Destroy()
		
		self.updateRecentFiles()

		# Restore the previous categories.
		race = Model.race
		importedCategories = False
		if categoriesFile:
			try:
				with open(categoriesFile, 'r') as fp:
					race.importCategories( fp )
				importedCategories = True
			except IOError:
				Utils.MessageOK( self, "Cannot open file:\n%s" % categoriesFile, "File Open Error", iconMask=wx.ICON_ERROR)
			except (ValueError, IndexError):
				Utils.MessageOK( self, "Bad file format:\n%s\n\n%s - %s" % (categoriesFile, str(ValueError), str(IndexError)), "File Read Error", iconMask=wx.ICON_ERROR)

		if not importedCategories:
			race.categories = categoriesSave

		if geoTrack:
			race.geoTrack, race.geoTrackFName = geoTrack, geoTrackFName
			distance = geoTrack.length if race.distanceUnit == race.UnitKm else geoTrack.length * 0.621371
			if distance > 0.0:
				for c in race.categories.itervalues():
					c.distance = distance
			race.showOval = False
		if excelLink:
			race.excelLink = excelLink
		
		self.setActiveCategories()
		self.setNumSelect( None )
		self.writeRace()
		self.showPageName( 'Actions' )
		self.refreshAll()

	def updateRecentFiles( self ):
		self.filehistory.AddFileToHistory(self.fileName)
		self.filehistory.Save(self.config)
		self.config.Flush()
		
	def closeFindDialog( self ):
		if getattr(self, 'findDialog', None):
			self.findDialog.Show( False )

	def openRace( self, fileName ):
		if not fileName:
			return
		self.showPageName( 'Results' )
		self.refresh()
		Model.resetCache()
		ResetExcelLinkCache()
		self.writeRace()
		self.closeFindDialog()
		
		try:
			with open(fileName, 'rb') as fp, Model.LockRace() as race:
				race = pickle.load( fp )
				race.sortLap = None			# Remove results lap sorting to avoid confusion.
				isFinished = race.isFinished()
				race.tagNums = None
				race.resetAllCaches()
				Model.setRace( race )
			
			self.fileName = fileName
			
			undo.clear()
			ResetExcelLinkCache()
			ResetPhotoInfoCache( self.fileName )
			Model.resetCache()
			
			self.updateRecentFiles()
			
			self.setNumSelect( None )
			self.record.setTimeTrialInput( getattr(race, 'isTimeTrial', False) )
			self.showPageName( 'Results' if isFinished else 'Actions')
			self.refreshAll()
			Utils.writeLog( 'call: openRace: "%s"' % fileName )

		except IOError:
			Utils.MessageOK(self, 'Cannot open file "%s".' % fileName, 'Cannot Open File', iconMask=wx.ICON_ERROR )

	@logCall
	def menuOpen( self, event ):
		dlg = wx.FileDialog( self, message="Choose a Race file",
							defaultFile = '',
							wildcard = 'CrossMan files (*.cmn)|*.cmn',
							style=wx.OPEN | wx.CHANGE_DIR )
		if dlg.ShowModal() == wx.ID_OK:
			self.openRace( dlg.GetPath() )
		dlg.Destroy()

	def menuFileHistory( self, event ):
		fileNum = event.GetId() - wx.ID_FILE1
		fileName = self.filehistory.GetHistoryFile(fileNum)
		self.filehistory.AddFileToHistory(fileName)  # move up the list
		self.openRace( fileName )
		
	@logCall
	def menuOpenNext( self, event ):
		race = Model.race

		if race is None or not self.fileName:
			self.menuOpen( event )
			return

		if race is not None and race.isRunning():
			if not Utils.MessageOKCancel(self,	'The current race is still running.\nFinish it and continue?',
												'Current race running' ):
				return
			race.finishRaceNow()
			self.writeRace()

		path, file = os.path.split( self.fileName )
		patRE = re.compile( '\-r[0-9]+\-' )
		try:
			pos = patRE.search(file).start()
			prefix = file[:pos+2] + str(race.raceNum+1)
			fileName = (f for f in os.listdir(path) if f.startswith(prefix)).next()
			fileName = os.path.join( path, fileName )
			self.openRace( fileName )
		except (AttributeError, IndexError, StopIteration):
			Utils.MessageOK(self, 'No next race.', 'No next race', iconMask=wx.ICON_ERROR )

	@logCall
	def menuExit(self, event):
		self.onCloseWindow( event )

	def genTimes( self ):
		regen = False
		if regen:
			fname = 'SimulationLapTimes.py'
		
			# Generate all the random rider events.
			random.seed( 10101021 )

			self.raceMinutes = 8
			mean = 8*60.0 / 8	# Average lap time.
			var = mean/20		# Variance between riders.
			lapsTotal = int(self.raceMinutes * 60 / mean + 3)
			raceTime = mean * lapsTotal
			errorPercent = 1.0/25.0

			riders = 30
			numStart = 200 - riders/2

			self.lapTimes = []
			for num in xrange(numStart,numStart+riders+1):
				t = 0
				if num < numStart + riders / 2:
					mu = random.normalvariate( mean, mean/20.0 )			# Rider's random average lap time.
				else:
					mu = random.normalvariate( mean * 1.15, mean/20.0 )		# These riders are slower, on average.
					t += 60.0												# Account for offset start.
				for laps in xrange(lapsTotal):
					t += random.normalvariate( mu, var/2 )	# Rider's lap time.
					if random.random() > errorPercent:		# Respect error rate.
						self.lapTimes.append( (t, num) )

			self.lapTimes.sort()
			
			# Get the times and leaders for each lap.
			leaderTimes = [self.lapTimes[0][0]]
			leaderNums  = [self.lapTimes[0][1]]
			numSeen = set()
			for t, n in self.lapTimes:
				if n in numSeen:
					leaderTimes.append( t )
					leaderNums.append( n )
					numSeen.clear()
				numSeen.add( n )
			
			# Find the leader's time after the end of the race.
			iLast = bisect.bisect_left( leaderTimes, self.raceMinutes * 60.0, hi = len(leaderTimes) - 1 )
			if leaderTimes[iLast] < self.raceMinutes * 60.0:
				iLast += 1
				
			# Trim out everything except next arrivals after the finish time.
			tLeaderLast = leaderTimes[iLast]
			numSeen = set()
			afterLeaderFinishEvents = [evt for evt in self.lapTimes if evt[0] >= tLeaderLast]
			self.lapTimes = [evt for evt in self.lapTimes if evt[0] < tLeaderLast]
			
			# Find the next unique arrival of all finishers.
			lastLapFinishers = []
			tStop = self.raceMinutes * 60.0
			numSeen = set()
			for t, n in afterLeaderFinishEvents:
				if n not in numSeen:
					numSeen.add( n )
					lastLapFinishers.append( (t, n) )
					
			self.lapTimes.extend( lastLapFinishers )
			
			with open(fname, 'w') as f:
				print >> f, 'raceMinutes =', self.raceMinutes
				print >> f, 'lapTimes =', self.lapTimes
		else:
			self.raceMinutes = SimulationLapTimes.raceMinutes
			self.lapTimes = copy.copy(SimulationLapTimes.lapTimes)
			
		return self.lapTimes
		
	@logCall
	def menuSimulate( self, event ):
		fName = os.path.join( Utils.getHomeDir(), 'Simulation.cmn' )
		if not Utils.MessageOKCancel(self,
'''
This will simulate a race using randomly generated data.
It is a good illustration of CrossMgr's functionality with realtime data.

The simulation takes about 8 minutes.

Unlike a real race, the simulation will show riders crossing the line right from the start.  In a real race, riders would have to finish the first lap before they were recorded.

The race will be written to:
"%s".

Continue?''' % fName, 'Simulate a Race' ):
			return

		try:
			with open(fName, 'wb') as fp:
				pass
		except IOError:
			Utils.MessageOK(self, 'Cannot open file "%s".' % fName, 'File Open Error',iconMask = wx.ICON_ERROR)
			return

		self.showPageName( 'Results' )	# Switch to a read-only view.
		self.refresh()
		
		self.lapTimes = self.genTimes()
		tMin = self.lapTimes[0][0]
		self.lapTimes.reverse()			# Reverse the times so we can pop them from the stack later.

		# Set up a new file and model for the simulation.
		self.fileName = fName
		OutputStreamer.DeleteStreamerFile()
		self.simulateSeen = set()
		undo.clear()
		with Model.lock:
			Model.setRace( None )
			race = Model.newRace()
			race.name = 'Simulate'
			race.organizer = 'Edward Sitarski'
			race.memo = ''
			race.minutes = self.raceMinutes
			race.raceNum = 1
			#race.isTimeTrial = True
			race.enableUSBCamera = True
			race.enableJChipIntegration = True
			race.setCategories( [	{'name':'Junior', 'catStr':'100-199', 'startOffset':'00:00', 'distance':0.5, 'gender':'Men'},
									{'name':'Senior', 'catStr':'200-299', 'startOffset':'00:15', 'distance':0.5, 'gender':'Women'}] )

		self.writeRace()
		DeletePhotos( self.fileName )
		
		# Create an Excel data file of rider data.
		fnameInfo = os.path.join( Utils.getImageFolder(), 'NamesTeams.csv' )
		riderInfo = []
		try:
			with open(fnameInfo) as fp:
				header = None
				for line in fp:
					line = line.decode('iso-8859-1')
					if not header:
						header = line.split(',')
						continue
					riderInfo.append( line.split(',') )
		except IOError:
			pass
			
		if riderInfo:
			wb = openpyxl.workbook.Workbook()
			ws = wb.create_sheet()
			ws.title = 'RiderData'
			for c, h in enumerate(['Bib#', 'LastName', 'FirstName', 'Team']):
				ws.cell(row = 0, column = c).value = h
			for r, row in enumerate(riderInfo):
				ws.cell( row = r + 1, column = 0 ).value = r + 100
				for c, v in enumerate(row):
					ws.cell( row = r + 1, column = c + 1 ).value = v
			fnameRiderInfo = os.path.join(Utils.getHomeDir(), 'SimulationRiderData.xlsx')
			wb.save( fnameRiderInfo )
			race.excelLink = ExcelLink()
			race.excelLink.setFileName( fnameRiderInfo )
			race.excelLink.setSheetName( ws.title )
			race.excelLink.setFieldCol( {'Bib#':0, 'LastName':1, 'FirstName':2, 'Team':3} )

		# Start the simulation.
		self.showPageName( 'History' )
		self.refresh()

		self.nextNum = None
		with Model.LockRace() as race:
			race.startRaceNow()		
			# Backup all the events and race start so we don't have to wait for the first lap.
			race.startTime -= datetime.timedelta( seconds = (tMin-1) )

		OutputStreamer.writeRaceStart()
		self.simulateTimer = wx.CallLater( 1, self.updateSimulation, True )
		self.updateRaceClock()
		self.refresh()

	def updateSimulation( self, num ):
		if Model.race is None:
			return
		if self.nextNum is not None and self.nextNum not in self.simulateSeen:
			self.forecastHistory.logNum( self.nextNum )

		with Model.LockRace() as race:
			if race.curRaceTime() > race.minutes * 60.0:
				self.simulateSeen.add( self.nextNum )

		try:
			t, self.nextNum = self.lapTimes.pop()
			with Model.LockRace() as race:
				if t < (self.raceMinutes*60.0 + race.getAverageLapTime()*1.5):
					self.simulateTimer.Restart( int(max(1,(t - race.curRaceTime()) * 1000)), True )
					return
		except IndexError:
			pass
			
		self.simulateTimer.Stop()
		self.nextNum = None
		with Model.LockRace() as race:
			race.finishRaceNow()
		VideoBuffer.Shutdown()
		SetCameraState( False )
		JChip.Cleanuplistener()
		
		OutputStreamer.writeRaceFinish()
		# Give the streamer a chance to write the last message.
		wx.CallLater( 2000, OutputStreamer.StopStreamer )
			
		Utils.writeRace()
		self.refresh()

	@logCall
	def menuImportCategories( self, event ):
		self.commit()
		if not Model.race:
			Utils.MessageOK( self, "A race must be loaded first.", "Import Categories", iconMask=wx.ICON_ERROR)
			return
			
		dlg = wx.FileDialog( self, message="Choose Race Categories File",
							defaultDir=os.getcwd(), 
							defaultFile="",
							wildcard="Bicycle Race Categories (*.brc)|*.brc",
							style=wx.OPEN )
		if dlg.ShowModal() == wx.ID_OK:
			categoriesFile = dlg.GetPath()
			try:
				with open(categoriesFile, 'r') as fp, Model.LockRace() as race:
					race.importCategories( fp )
			except IOError:
				Utils.MessageOK( self, "Cannot open file:\n%s" % categoriesFile, "File Open Error", iconMask=wx.ICON_ERROR)
			except (ValueError, IndexError):
				Utils.MessageOK( self, "Bad file format:\n%s" % categoriesFile, "File Read Error", iconMask=wx.ICON_ERROR)
				
		dlg.Destroy()
	
	@logCall
	def menuExportCategories( self, event ):
		self.commit()
		race = Model.getRace()
		if not race:
			Utils.MessageOK( self, "A race must be loaded first.", "Import Categories", iconMask=wx.ICON_ERROR)
			return
			
		dlg = wx.FileDialog( self, message="Choose Race Categories File",
							defaultDir=os.getcwd(), 
							defaultFile="",
							wildcard="Bicycle Race Categories (*.brc)|*.brc",
							style=wx.SAVE )
							
		if dlg.ShowModal() == wx.ID_OK:
			fname = dlg.GetPath()
			try:
				with open(fname, 'w') as fp, Model.LockRace() as race:
					race.exportCategories( fp )
			except IOError:
				Utils.MessageOK( self, "Cannot open file:\n%s" % fname, "File Open Error", iconMask=wx.ICON_ERROR)
				
		dlg.Destroy()	
		
	@logCall
	def menuExportHistory( self, event ):
		self.commit()
		if self.fileName is None or len(self.fileName) < 4 or not Model.race:
			return

		self.showPageName( 'History' )
		self.history.setCategoryAll()
		self.history.refresh()
		
		xlFName = self.fileName[:-4] + '-History.xls'
		dlg = wx.DirDialog( self, 'Folder to write "%s"' % os.path.basename(xlFName),
						style=wx.DD_DEFAULT_STYLE, defaultPath=os.path.dirname(xlFName) )
		ret = dlg.ShowModal()
		dName = dlg.GetPath()
		dlg.Destroy()
		if ret != wx.ID_OK:
			return

		xlFName = os.path.join( dName, os.path.basename(xlFName) )

		colnames = self.history.grid.GetColNames()
		data = self.history.grid.GetData()
		if data:
			rowMax = max( len(c) for c in data )
			colnames = ['Count'] + colnames
			data = [[str(i) for i in xrange(1, rowMax+1)]] + data
		with Model.LockRace() as race:
			title = 'Race: '+ race.name + '\n' + Utils.formatDate(race.date) + '\nRace History'
		export = ExportGrid( title, colnames, data )

		wb = xlwt.Workbook()
		sheetCur = wb.add_sheet( 'History' )
		export.toExcelSheet( sheetCur )
		
		try:
			wb.save( xlFName )
			webbrowser.open( xlFName, new = 2, autoraise = True )
			Utils.MessageOK(self, 'Excel file written to:\n\n   %s' % xlFName, 'Excel Write')
		except IOError:
			Utils.MessageOK(self,
						'Cannot write "%s".\n\nCheck if this spreadsheet is open.\nIf so, close it, and try again.' % xlFName,
						'Excel File Error', iconMask=wx.ICON_ERROR )
	
	@logCall
	def menuExportUSAC( self, event ):
		self.commit()
		if self.fileName is None or len(self.fileName) < 4 or not Model.race:
			return

		self.showPageName( 'Results' )
		
		xlFName = self.fileName[:-4] + '-USAC.xls'
		dlg = wx.DirDialog( self, 'Folder to write "%s"' % os.path.basename(xlFName),
						style=wx.DD_DEFAULT_STYLE, defaultPath=os.path.dirname(xlFName) )
		ret = dlg.ShowModal()
		dName = dlg.GetPath()
		dlg.Destroy()
		if ret != wx.ID_OK:
			return

		xlFName = os.path.join( dName, os.path.basename(xlFName) )

		wb = xlwt.Workbook()
		sheetCur = wb.add_sheet( 'Combined Results' )
		USACExport( sheetCur )
		
		try:
			wb.save( xlFName )
			webbrowser.open( xlFName, new = 2, autoraise = True )
			Utils.MessageOK(self, 'Excel file written to:\n\n   %s' % xlFName, 'Excel Write')
		except IOError:
			Utils.MessageOK(self,
						'Cannot write "%s".\n\nCheck if this spreadsheet is open.\nIf so, close it, and try again.' % xlFName,
						'Excel File Error', iconMask=wx.ICON_ERROR )
	
	@logCall
	def menuHelpQuickStart( self, event ):
		Utils.showHelp( 'QuickStart.html' )
	
	@logCall
	def menuHelpSearch( self, event ):
		self.helpSearch.Show()
	
	@logCall
	def menuHelp( self, event ):
		Utils.showHelp( 'Main.html' )
	
	@logCall
	def onContextHelp( self, event ):
		Utils.showHelp( self.attrClassName[self.notebook.GetSelection()][2] + '.html' )
	
	@logCall
	def menuAbout( self, event ):
		# First we create and fill the info object
		info = wx.AboutDialogInfo()
		info.Name = Version.AppVerName
		info.Version = ''
		info.Copyright = "(C) 2009-2012"
		info.Description = wordwrap(
			"Create Cyclo-cross race results quickly and easily with little preparation.\n\n"
			"A brief list of features:\n"
			"   * Input riders on the first lap\n"
			"   * Predicts riders for all other laps based on lap times\n"
			"   * Indicates race leader by category and tracks missing riders\n"
			"   * Interpolates missing numbers.  Ignores duplicate rider entries.\n"
			"   * Shows results instantly by category during and after race\n"
			"   * Shows rider history\n"
			"   * Allows rider lap adjustments\n"
			"   * UCI 80% Rule Countdown\n"
			"",
			500, wx.ClientDC(self))
		info.WebSite = ("http://sites.google.com/site/crossmgrsoftware/", "CrossMgr Home Page")
		info.Developers = [
					"Edward Sitarski (edward.sitarski@gmail.com)",
					"Andrew Paradowski (andrew.paradowski@gmail.com)"
					]

		licenseText = "User Beware!\n\n" \
			"This program is experimental, under development and may have bugs.\n" \
			"Feedback is sincerely appreciated.\n\n" \
			"Donations are also appreciated - see website for details.\n\n" \
			"CRITICALLY IMPORTANT MESSAGE!\n" \
			"This program is not warrented for any use whatsoever.\n" \
			"It may not produce correct results, it might lose your data.\n" \
			"The authors of this program assume no reponsibility or liability for data loss or erronious results produced by this program.\n\n" \
			"Use entirely at your own risk.\n" \
			"Do not come back and tell me that this program screwed up your event!\n" \
			"Computers fail, screw-ups happen.  Always use a paper manual backup."
		info.License = wordwrap(licenseText, 500, wx.ClientDC(self))

		wx.AboutBox(info)

	#--------------------------------------------------------------------------------------

	def getCurrentPage( self ):
		return self.pages[self.notebook.GetSelection()]
	
	def showRiderDetail( self, num = None ):
		self.riderDetail.setRider( num )
		for i, p in enumerate(self.pages):
			if p == self.riderDetail:
				self.showPage( i )
				break

	def setRiderDetail( self, num = None ):
		self.riderDetail.setRider( num )

	def showPage( self, iPage ):
		self.callPageCommit( self.notebook.GetSelection() )
		self.callPageRefresh( iPage )
		self.notebook.ChangeSelection( iPage )
		self.pages[self.notebook.GetSelection()].Layout()

	def showPageName( self, name ):
		name = name.replace(' ', '')
		for i, (a, c, n) in enumerate(self.attrClassName):
			if n == name:
				self.showPage( i )
				break

	def showRiderDetail( self ):
		self.showPageName( 'RiderDetail' )
				
	def callPageRefresh( self, i ):
		try:
			page = self.pages[i]
		except IndexError:
			return
			
		if hasattr(page, 'refresh'):
			page.refresh()

	def callPageCommit( self, i ):
		try:
			self.pages[i].commit()
		except (AttributeError, IndexError):
			pass

	def commit( self ):
		self.callPageCommit( self.notebook.GetSelection() )
				
	def refreshCurrentPage( self ):
		self.callPageRefresh( self.notebook.GetSelection() )

	def refresh( self ):
		self.refreshCurrentPage()
		self.forecastHistory.refresh()
		if self.riderDetailDialog:
			wx.CallAfter( self.riderDetailDialog.refresh )
		self.updateRaceClock()
		with Model.LockRace() as race:
			self.menuItemHighPrecisionTimes.Check( bool(race and getattr(race, 'highPrecisionTimes', False)) ) 
			self.menuItemSyncCategories.Check( bool(race and getattr(race, 'syncCategories', True)) ) 
		if self.photoDialog.IsShown():
			self.photoDialog.refresh()

	def updateUndoStatus( self, event = None ):
		with Model.LockRace() as race:
			isRunning = bool(race and race.isRunning())
		self.undoMenuButton.Enable( bool(not isRunning and undo.isUndo()) )
		self.redoMenuButton.Enable( bool(not isRunning and undo.isRedo()) )
		
	def onPageChanging( self, event ):
		self.callPageCommit( event.GetOldSelection() )
		self.callPageRefresh( event.GetSelection() )
		try:
			Utils.writeLog( 'page: %s\n' % self.pages[event.GetSelection()].__class__.__name__ )
		except IndexError:
			pass
		event.Skip()	# Required to properly repaint the screen.

	def refreshRaceAnimation( self ):
		if self.pages[self.notebook.GetSelection()] == self.raceAnimation:
			self.raceAnimation.refresh()
	
	def refreshAll( self ):
		self.refresh()
		iSelect = self.notebook.GetSelection()
		for i, p in enumerate(self.pages):
			if i != iSelect:
				self.callPageRefresh( i )

	def setNumSelect( self, num ):
		try:
			num = int(num)
		except (TypeError, ValueError):
			num = None
			
		if num is None or num != self.numSelect:
			self.history.setNumSelect( num )
			self.results.setNumSelect( num )
			self.riderDetail.setNumSelect( num )
			self.gantt.setNumSelect( num )
			self.raceAnimation.setNumSelect( num )
			if self.photoDialog.IsShown():
				self.photoDialog.setNumSelect( num )
			self.numSelect = num

	#-------------------------------------------------------------
	
	def processJChipListener( self ):
		
		# Assumes Model is locked.
		race = Model.race
		
		if not race or not getattr(race, 'enableJChipIntegration', False):
			if JChip.listener:
				JChip.StopListener()
			return False
		
		if not JChip.listener:
			JChip.StartListener( race.startTime )
			JChipSetup.GetTagNums( True )
		
		data = JChip.GetData()
		
		if not getattr(race, 'tagNums', None):
			JChipSetup.GetTagNums()
		if not race.tagNums:
			return False
		
		success = False
		numTimes = []
		for d in data:
			if d[0] != 'data':
				continue
			try:
				num, dt = race.tagNums[d[1]], d[2]
			except (TypeError, ValueError, KeyError):
				race.missingTags.add( d[1] )
				continue
				
			# Ignore times before the start of the race.
			if race.isRunning() and race.startTime <= dt:
				delta = dt - race.startTime
				t = delta.total_seconds()
				t = race.addTime( num, t )
				numTimes.append( (num, t) )
				success = True
		
		if success:
			OutputStreamer.writeNumTimes( numTimes )
			if self.getCurrentPage() == self.results:
				wx.CallAfter( self.results.showLastLap )
			if getattr(race, 'ftpUploadDuringRace', False):
				realTimeFtpPublish.publishEntry()

		return success

	def updateRaceClock( self, event = None ):
		if hasattr(self, 'record'):
			self.record.refreshRaceTime()

		doRefresh = False
		with Model.LockRace() as race:
			if race is None:
				self.SetTitle( Version.AppVerName )
				self.timer.Stop()
				JChip.StopListener()
				return

			if race.isUnstarted():
				status = 'Unstarted'
			elif race.isRunning():
				status = 'Running'
				if getattr(race, 'enableJChipIntegration', False):
					doRefresh = self.processJChipListener()
				elif JChip.listener:
					JChip.StopListener()
			else:
				status = 'Finished'

			if not race.isRunning():
				self.SetTitle( '%s-r%d - %s - %s %s' % (
								race.name, race.raceNum,
								status,
								Version.AppVerName,
								'<TimeTrial>' if getattr(race, 'isTimeTrial', False) else '') )
				self.timer.Stop()
				return

			self.SetTitle( '%s %s-r%d - %s - %s %s %s' % (
							Utils.formatTime(race.curRaceTime()),
							race.name, race.raceNum,
							status, Version.AppVerName,
							'<JChip>' if JChip.listener else '',
							'<TimeTrial>' if getattr(race, 'isTimeTrial', False) else '') )

		if self.timer is None or not self.timer.IsRunning():
			self.timer.Start( 1000 )
			self.secondCount = 0

		self.secondCount += 1
		if self.secondCount % 30 == 0 and race.isChanged():
			self.writeRace()
			
		if doRefresh:
			wx.CallAfter( self.refresh )
			wx.CallAfter( self.record.refreshLaps )

# Set log file location.
dataDir = ''
redirectFileName = ''
			
def MainLoop():
	global dataDir
	global redirectFileName
	
	random.seed()
	setpriority( priority=4 )	# Set to real-time priority.

	parser = OptionParser( usage = "usage: %prog [options] [RaceFile.cmn]" )
	parser.add_option("-f", "--file", dest="filename", help="race file", metavar="RaceFile.cmn")
	parser.add_option("-q", "--quiet", action="store_false", dest="verbose", default=True, help='hide splash screen')
	parser.add_option("-r", "--regular", action="store_false", dest="fullScreen", default=True, help='regular size (not full screen)')
	(options, args) = parser.parse_args()

	app = wx.PySimpleApp()
	app.SetAppName("CrossMgr")
	
	dataDir = Utils.getHomeDir()
	redirectFileName = os.path.join(dataDir, 'CrossMgr.log')
			
	# Set up the log file.  Otherwise, show errors on the screen unbuffered.
	if __name__ == '__main__':
		Utils.disable_stdout_buffering()
	else:
		try:
			logSize = os.path.getsize( redirectFileName )
			if logSize > 1000000:
				os.remove( redirectFileName )
		except:
			pass
	
		try:
			app.RedirectStdio( redirectFileName )
		except:
			pass
	
	Utils.writeLog( 'start: %s' % Version.AppVerName )
	
	# Configure the main window.
	mainWin = MainWin( None, title=Version.AppVerName, size=(1128,600) )
	if options.fullScreen:
		mainWin.Maximize( True )
	mainWin.Show()

	# Set the upper left icon.
	icon = wx.Icon( os.path.join(Utils.getImageFolder(), 'CrossMgr16x16.ico'), wx.BITMAP_TYPE_ICO )
	mainWin.SetIcon( icon )

	# Set the taskbar icon.
	#tbicon = wx.TaskBarIcon()
	#tbicon.SetIcon( icon, "CrossMgr" )

	if options.verbose:
		ShowSplashScreen()
		ShowTipAtStartup()
		# multiprocessing.Process( target = VersionMgr.updateVersionCache, args = (VersionMgr.getVersionFileName(),)).start()
		# threading.Thread( target = VersionMgr.updateVersionCache ).run()
	
	# Try to open a specified filename.
	fileName = options.filename
	
	# If nothing, try a positional argument.
	if not fileName and args:
		fileName = args[0]
	
	# Try to load a race.
	if fileName:
		try:
			mainWin.openRace( fileName )
		except (IndexError, AttributeError, ValueError):
			pass

	mainWin.forecastHistory.setSash()
	
	# Start processing events.
	app.MainLoop()

if __name__ == '__main__':
	MainLoop()
