#!/usr/bin/env python
#------------------------------------------------------------------------------	
# JChipClient.py: JChip simulator program for testing JChip interface and CrossMgr.
#
# Copyright (C) Edward Sitarski, 2012.
import re
import os
import sys
import time
import socket
import random
import datetime
import subprocess
from openpyxl.workbook import Workbook

#------------------------------------------------------------------------------	
# CrossMgr's port and socket.
DEFAULT_PORT = 53135
DEFAULT_HOST = '127.0.0.1'

#------------------------------------------------------------------------------	
# JChip delimiter (CR, **not** LF)
CR = chr( 0x0d )

#------------------------------------------------------------------------------	
# Create some random rider numbers.
random.seed( 10101010 )
seen = set()
nums = []
for i in xrange(25):
	while 1:
		x = random.randint(1,200)
		if x not in seen:
			seen.add( x )
			nums.append( x )
			break

#------------------------------------------------------------------------------	
# Create a JChip-style hex tag for each number.
tag = dict( (n, '413A%02X' % n) for n in nums )
tag[random.choice(list(tag.keys()))] = 'E2001018860B01290700D0D8'
tag[random.choice(list(tag.keys()))] = 'E2001018860B01530700D138'
tag[random.choice(list(tag.keys()))] = 'E2001018860B01370700D0F8'

#------------------------------------------------------------------------------	
# Write out as a .xlsx file with the number tag data.
#
wb = Workbook()
ws = wb.get_active_sheet()
ws.title = "JChipTest"
for col, label in enumerate('Bib#,Tag'.split(',')):
	ws.cell( row = 0, column = col ).value = label
for row, n in enumerate(nums):
	ws.cell( row = row + 1, column = 0 ).value = n
	ws.cell( row = row + 1, column = 1 ).value = tag[n]
wb.save('JChipTest.xlsx')
wb = None

#------------------------------------------------------------------------------	
# Also write out as a .csv file.
#
with open('JChipTest.csv', 'w') as f:
	f.write( 'Bib#,Tag,dummy3,dummy4,dummy5\n' )
	for n in nums:
		f.write( '%d,%s\n' % (n, tag[n]) )

sendDate = True

#------------------------------------------------------------------------------	
# Function to format number, lap and time in JChip format
# Z413A35 10:11:16.4433 10  10000      C7
count = 0
def formatMessage( n, lap, t ):
	global count
	message = "DJ%s %s 10  %05X      C7%s%s" % (
				tag[n],								# Tag code
				t.strftime('%H:%M:%S.%f'),			# hh:mm:ss.ff
				count,								# Data index number in hex.
				' date={}'.format( t.strftime('%Y%m%d') ) if sendDate else '',
				CR
			)
	count += 1
	return message

#------------------------------------------------------------------------------	
# Generate some random lap times.
random.seed()
numLapTimes = []
mean = 60.0							# Average lap time.
varFactor = 9.0 * 4.0
var = mean / varFactor				# Variance between riders.
lapMax = 6
for n in nums:
	lapTime = random.normalvariate( mean, mean/(varFactor * 4.0) )
	for lap in xrange(0, lapMax+1):
		numLapTimes.append( (n, lap, lapTime*lap) )
numLapTimes.sort( key = lambda x: (x[1], x[2]) )	# Sort by lap, then race time.

#------------------------------------------------------------------------------	
# Create a socket (SOCK_STREAM means a TCP socket)
sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

#------------------------------------------------------------------------------	
# Connect to the CrossMgr server.
iMessage = 1
while 1:
	print 'Trying to connect to server...'
	while 1:
		try:
			sock.connect((DEFAULT_HOST, DEFAULT_PORT))
			break
		except:
			print 'Connection failed.  Waiting 5 seconds...'
			time.sleep( 5 )

	#------------------------------------------------------------------------------	
	print 'Connection succeeded!'
	name = '{}-{}'.format(socket.gethostname(), os.getpid())
	print 'Sending name...', name
	sock.send( "N0000{}{}".format(name, CR) )

	#------------------------------------------------------------------------------	
	print 'Waiting for get time command...'
	while 1:
		received = sock.recv(1)
		if received == 'G':
			while received[-1] != CR:
				received += sock.recv(1)
			print 'Received cmd: "%s" from CrossMgr' % received[:-1]
			break

	#------------------------------------------------------------------------------	
	dBase = datetime.datetime.now()
	dBase -= datetime.timedelta( seconds = 13*60+13.13 )	# Send the wrong time for testing purposes.

	#------------------------------------------------------------------------------	
	print 'Send gettime data...'
	# format is GT0HHMMSShh<CR> where hh is 100's of a second.  The '0' (zero) after GT is the number of days running and is ignored by CrossMgr.
	message = 'GT0%02d%02d%02d%02d%s%s' % (
		dBase.hour, dBase.minute, dBase.second, int((dBase.microsecond / 1000000.0) * 100.0),
		' date={}'.format( dBase.strftime('%Y%m%d') ) if sendDate else '',
		CR)
	print message[:-1]
	sock.send( message )

	#------------------------------------------------------------------------------	
	print 'Waiting for send command from CrossMgr...'
	while 1:
		received = sock.recv(1)
		if received == 'S':
			while received[-1] != CR:
				received += sock.recv(1)
			print 'Received cmd: "%s" from CrossMgr' % received[:-1]
			break

	#------------------------------------------------------------------------------	
	print 'Start sending data...'

	while iMessage < len(numLapTimes):
		n, lap, t = numLapTimes[iMessage]
		dt = t - numLapTimes[iMessage-1][2]
		
		time.sleep( dt )
		
		message = formatMessage( n, lap, dBase + datetime.timedelta(seconds = t - 0.5) )
		sys.stdout.write( 'sending: %d: %s\n' % (iMessage, message[:-1]) )
		try:
			sock.send( message )
			iMessage += 1
		except:
			print 'Disconnected.  Attempting to reconnect...'
			break
		
		
	if iMessage >= len(numLapTimes):
		sock.send( '<<<GarbageTerminateMessage>>>' + CR )
		break
		
